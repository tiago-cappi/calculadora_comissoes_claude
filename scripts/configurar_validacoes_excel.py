"""
configurar_validacoes_excel.py — Aplica dropdowns com filtros BIDIRECIONAIS
(dinâmicos) ao `configuracoes_comissoes.xlsx`.

Requer Excel 365 / Excel 2021 ou superior (fórmulas de array dinâmico:
FILTER, UNIQUE, SORT) e Windows com Excel instalado (usa COM para finalizar
a serialização, já que openpyxl não marca corretamente células com spill).

Filtros configurados
--------------------
- Abas `config_comissao`, `metas_aplicacao`, `meta_rentabilidade`:
  qualquer valor selecionado em `linha`, `grupo`, `subgrupo`,
  `tipo_mercadoria` ou `fabricante` filtra automaticamente os dropdowns
  das outras 4 colunas da mesma linha, em qualquer ordem.
- Abas `config_comissao`, `pesos_metas`, `metas_individuais`:
  `cargo` e `colaborador` se filtram mutuamente.
- Dropdowns simples: `cross_selling.colaborador`,
  `metas_fornecedores.linha`, `pesos_metas.linha`.

Como funciona
-------------
Para cada célula-alvo de dropdown bidirecional, o script mantém uma célula
auxiliar oculta com uma fórmula SPILL (horizontal, via TRANSPOSE) que
calcula os valores compatíveis com as outras seleções da mesma linha. A
validação de dados do Excel aponta para o spill via referência `#`
(ex.: `=$Z2#`), padrão que o Excel 365 reconhece nativamente em DV.

Uso
---
    python scripts/configurar_validacoes_excel.py
    python scripts/configurar_validacoes_excel.py --input outro.xlsx

Idempotente: pode ser executado múltiplas vezes. Sempre cria backup
datado antes de sobrescrever o arquivo original.
"""

from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName


PRE_ALLOC_ROWS = 200  # pré-aloca DV e spills até a linha 201
HIERARCHY_DIMS = ["linha", "grupo", "subgrupo", "tipo_mercadoria", "fabricante"]
HIERARCHY_TARGETS = ["config_comissao", "metas_aplicacao", "meta_rentabilidade"]
CARGO_COLAB_TARGETS = ["config_comissao", "pesos_metas", "metas_individuais"]

# Índices (1-based) das colunas auxiliares ocultas. Espaçamento de 500 colunas
# entre cada dimensão, suficiente para spills horizontais das maiores listas
# (subgrupo pode ter até ~300 valores únicos).
HELPER_COL_IDX: Dict[str, int] = {
    "linha":           30,     # AD
    "grupo":           530,
    "subgrupo":        1030,
    "tipo_mercadoria": 1530,
    "fabricante":      2030,
    "cargo":           2530,
    "colaborador":     3030,
}

DV_ERROR_TITLE = "Valor fora da lista"
DV_ERROR_MSG = (
    "O valor digitado não consta na configuração. Se for um cadastro novo, "
    "adicione em 'colaboradores' / 'cargos' / 'classificacao_produtos' antes. "
    "Caso contrário, corrija para evitar erros no cálculo."
)

# Constantes do Excel COM
_XL_VALIDATE_LIST = 3
_XL_VALID_ALERT_STOP = 1
_XL_VALID_ALERT_WARNING = 2
_XL_BETWEEN = 1


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def _backup(path: Path) -> Path:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = path.with_name(f"{path.stem}.bak.{stamp}{path.suffix}")
    shutil.copy2(path, backup_path)
    return backup_path


def _header_map_openpyxl(wb, sheet: str) -> Dict[str, str]:
    ws = wb[sheet]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return {
        str(h).strip(): get_column_letter(i + 1)
        for i, h in enumerate(header)
        if h is not None
    }


def _header_map_com(ws) -> Dict[str, str]:
    headers: Dict[str, str] = {}
    col = 1
    while True:
        v = ws.Cells(1, col).Value
        if v is None or v == "":
            break
        headers[str(v).strip()] = get_column_letter(col)
        col += 1
    return headers


# ---------------------------------------------------------------------------
# Etapa 1 — openpyxl: limpar estado anterior e registrar named ranges
# ---------------------------------------------------------------------------

def _prepare_with_openpyxl(path: Path) -> None:
    wb = load_workbook(path)

    # Remove sheets legadas de versões anteriores do script
    for legacy in ("_dv_wf", "_dv_hierarquia", "_dv_helpers"):
        if legacy in wb.sheetnames:
            del wb[legacy]

    # Remove tabelas nomeadas em qualquer sheet (residuais)
    for ws in wb.worksheets:
        if hasattr(ws, "tables"):
            for tn in list(ws.tables.keys()):
                del ws.tables[tn]

    # Remove named ranges antigas com prefixos reservados
    for name in list(wb.defined_names):
        if any(name.startswith(p) for p in ("dv_", "cp_", "cb_", "cg_")):
            del wb.defined_names[name]

    # Limpa DVs anteriores em todas as abas-alvo
    target_sheets = (
        set(HIERARCHY_TARGETS)
        | set(CARGO_COLAB_TARGETS)
        | {"cross_selling", "metas_fornecedores"}
    )
    for sheet in target_sheets:
        if sheet in wb.sheetnames:
            ws = wb[sheet]
            if ws.data_validations:
                ws.data_validations.dataValidation = []

    # Named ranges dinâmicos (OFFSET + COUNTA)
    cp = _header_map_openpyxl(wb, "classificacao_produtos")
    missing = [d for d in HIERARCHY_DIMS if d not in cp]
    if missing:
        raise RuntimeError(
            f"classificacao_produtos não possui as colunas: {missing}"
        )
    anchor = cp["linha"]
    cp_len = f"COUNTA(classificacao_produtos!${anchor}:${anchor})-1"
    for dim in HIERARCHY_DIMS:
        letter = cp[dim]
        ref = (
            f"OFFSET(classificacao_produtos!${letter}$2,0,0,{cp_len},1)"
        )
        wb.defined_names[f"cp_{dim}"] = DefinedName(
            name=f"cp_{dim}", attr_text=ref
        )

    col = _header_map_openpyxl(wb, "colaboradores")
    cb_name_l = col["nome_colaborador"]
    cb_cargo_l = col["cargo"]
    cb_len = f"COUNTA(colaboradores!${cb_name_l}:${cb_name_l})-1"
    wb.defined_names["cb_nome"] = DefinedName(
        name="cb_nome",
        attr_text=f"OFFSET(colaboradores!${cb_name_l}$2,0,0,{cb_len},1)",
    )
    wb.defined_names["cb_cargo"] = DefinedName(
        name="cb_cargo",
        attr_text=f"OFFSET(colaboradores!${cb_cargo_l}$2,0,0,{cb_len},1)",
    )

    cg = _header_map_openpyxl(wb, "cargos")
    cg_name_l = cg["nome_cargo"]
    cg_len = f"COUNTA(cargos!${cg_name_l}:${cg_name_l})-1"
    wb.defined_names["cg_nome_cargo"] = DefinedName(
        name="cg_nome_cargo",
        attr_text=f"OFFSET(cargos!${cg_name_l}$2,0,0,{cg_len},1)",
    )

    wb.save(path)


# ---------------------------------------------------------------------------
# Etapa 2 — Excel COM: escrever helpers (spill) e adicionar DVs com `#`
# ---------------------------------------------------------------------------

def _hierarchy_spill_formula(dim: str, col_letters: Dict[str, str]) -> str:
    """Spill horizontal (via TRANSPOSE) dos valores únicos/ordenados de `dim`
    compatíveis com as seleções atuais das outras 4 dimensões da mesma linha.

    Cada condição é `($X2="") + (cp_dim=$X2)`: vale 1 se a coluna está vazia
    (ignora aquela condição) ou se há match. O produto das 4 condições é ≠0
    apenas onde todas batem.
    """
    conds = []
    for d in HIERARCHY_DIMS:
        if d == dim:
            continue
        c = col_letters[d]
        conds.append(f'(($' + c + '2="")+(cp_' + d + '=$' + c + '2))')
    condition = "*".join(conds)
    return (
        f'=IFERROR(TRANSPOSE(SORT(UNIQUE(FILTER(cp_{dim},{condition})))),"")'
    )


def _cargo_spill_formula(colab_cell: str) -> str:
    return (
        f'=IFERROR(TRANSPOSE(SORT(UNIQUE('
        f'IF({colab_cell}="",cg_nome_cargo,'
        f'FILTER(cb_cargo,cb_nome={colab_cell}))))),"")'
    )


def _colab_spill_formula(cargo_cell: str) -> str:
    return (
        f'=IFERROR(TRANSPOSE(SORT('
        f'IF({cargo_cell}="",cb_nome,'
        f'FILTER(cb_nome,cb_cargo={cargo_cell})))),"")'
    )


def _apply_dv(ws, target_range_str: str, formula1: str) -> None:
    r = ws.Range(target_range_str)
    r.Validation.Delete()
    r.Validation.Add(
        Type=_XL_VALIDATE_LIST,
        AlertStyle=_XL_VALID_ALERT_WARNING,
        Operator=_XL_BETWEEN,
        Formula1=formula1,
    )
    r.Validation.IgnoreBlank = True
    r.Validation.InCellDropdown = True
    r.Validation.ShowError = True
    r.Validation.ErrorTitle = DV_ERROR_TITLE
    r.Validation.ErrorMessage = DV_ERROR_MSG


def _finalize_with_com(path: Path) -> None:
    try:
        import win32com.client
    except ImportError as e:
        raise RuntimeError(
            "pywin32 não instalado — instale com `pip install pywin32`"
        ) from e

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(path))

        # Abas de hierarquia (5 dims, bidirecionais)
        for sheet in HIERARCHY_TARGETS:
            ws = wb.Sheets(sheet)
            hdr = _header_map_com(ws)
            col_letters = {d: hdr[d] for d in HIERARCHY_DIMS}

            for dim in HIERARCHY_DIMS:
                target_letter = hdr[dim]
                helper_letter = get_column_letter(HELPER_COL_IDX[dim])
                formula = _hierarchy_spill_formula(dim, col_letters)
                ws.Range(
                    f"{helper_letter}2:{helper_letter}{PRE_ALLOC_ROWS + 1}"
                ).Formula2 = formula
                ws.Columns(helper_letter).Hidden = True
                _apply_dv(
                    ws,
                    f"{target_letter}2:{target_letter}{PRE_ALLOC_ROWS + 1}",
                    f"=${helper_letter}2#",
                )
            print(f"    {sheet}: 5 dropdowns de cascata bidirecional")

        # cargo ↔ colaborador (bidirecional)
        for sheet in CARGO_COLAB_TARGETS:
            ws = wb.Sheets(sheet)
            hdr = _header_map_com(ws)
            if "cargo" not in hdr or "colaborador" not in hdr:
                continue
            cargo_letter = hdr["cargo"]
            colab_letter = hdr["colaborador"]
            cargo_helper = get_column_letter(HELPER_COL_IDX["cargo"])
            colab_helper = get_column_letter(HELPER_COL_IDX["colaborador"])

            ws.Range(
                f"{cargo_helper}2:{cargo_helper}{PRE_ALLOC_ROWS + 1}"
            ).Formula2 = _cargo_spill_formula(f"${colab_letter}2")
            ws.Columns(cargo_helper).Hidden = True
            _apply_dv(
                ws,
                f"{cargo_letter}2:{cargo_letter}{PRE_ALLOC_ROWS + 1}",
                f"=${cargo_helper}2#",
            )

            ws.Range(
                f"{colab_helper}2:{colab_helper}{PRE_ALLOC_ROWS + 1}"
            ).Formula2 = _colab_spill_formula(f"${cargo_letter}2")
            ws.Columns(colab_helper).Hidden = True
            _apply_dv(
                ws,
                f"{colab_letter}2:{colab_letter}{PRE_ALLOC_ROWS + 1}",
                f"=${colab_helper}2#",
            )
            print(f"    {sheet}: cargo <-> colaborador")

        # Dropdowns simples (sem filtragem cruzada)
        for sheet, col_name, named_range in [
            ("cross_selling", "colaborador", "cb_nome"),
            ("metas_fornecedores", "linha", "cp_linha"),
            ("pesos_metas", "linha", "cp_linha"),
        ]:
            ws = wb.Sheets(sheet)
            hdr = _header_map_com(ws)
            if col_name not in hdr:
                continue
            letter = hdr[col_name]
            _apply_dv(
                ws,
                f"{letter}2:{letter}{PRE_ALLOC_ROWS + 1}",
                f"={named_range}",
            )
            print(f"    {sheet}.{col_name}: dropdown simples")

        wb.Save()
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    default_path = Path(__file__).parent.parent / "configuracoes_comissoes.xlsx"
    parser.add_argument("--input", type=Path, default=default_path)
    args = parser.parse_args()

    path: Path = args.input.resolve()
    if not path.exists():
        print(f"[ERRO] Arquivo não encontrado: {path}")
        return 1

    print(f"Arquivo: {path}")
    backup = _backup(path)
    print(f"Backup: {backup.name}")

    print("  [1/2] openpyxl: limpando estado anterior e registrando named ranges")
    _prepare_with_openpyxl(path)

    print("  [2/2] Excel COM: escrevendo helpers de spill e data validations")
    _finalize_with_com(path)

    print("\nConcluído. Abra `configuracoes_comissoes.xlsx` e teste:")
    print("  1. Em config_comissao, selecione um fabricante: linha/grupo/subgrupo/tipo")
    print("     ficarão restritos automaticamente (em qualquer ordem).")
    print("  2. Em metas_individuais, selecione um colaborador: o cargo se ajusta.")
    print("Se novas linhas forem adicionadas em classificacao_produtos / colaboradores /")
    print("cargos, os named ranges OFFSET+COUNTA se auto-expandem — mas se você precisar")
    print(f"passar de {PRE_ALLOC_ROWS} linhas nas abas-alvo, rode o script novamente.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
