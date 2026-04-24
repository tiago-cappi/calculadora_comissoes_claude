"""
=============================================================================
Raiolanda BA — Relatório de Validação de Processos Faturados na Bahia
=============================================================================
Gera Excel de validação com processos FATURADOS onde UF=BA no mês,
para que a analista valide manualmente quais processos Raiolanda comissiona.

Saída: saida/MM_AAAA/validacao_ba_MM_AAAA.xlsx
=============================================================================
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd


def execute(
    df_analise_comercial: pd.DataFrame,
    mes: int,
    ano: int,
    output_dir: Optional[str] = None,
) -> dict:
    """Gera relatório Excel de validação de processos BA para Raiolanda.

    Args:
        df_analise_comercial: DataFrame da AC já filtrado pelo mês/ano.
        mes: Mês de apuração.
        ano: Ano de apuração.
        output_dir: Diretório de saída. Se None, usa saida/MM_AAAA/ relativo ao projeto.

    Returns:
        dict com 'ok', 'arquivo', 'num_processos', 'num_itens', 'erros'.
    """
    result: dict = {"ok": False, "arquivo": None, "num_processos": 0, "num_itens": 0, "erros": []}

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        result["erros"].append(f"openpyxl não disponível: {exc}")
        return result

    if df_analise_comercial is None or df_analise_comercial.empty:
        result["erros"].append("Análise Comercial vazia — sem dados para gerar relatório BA.")
        return result

    df = df_analise_comercial.copy()

    # Filtrar FATURADO e UF=BA
    status_col = "Status Processo"
    uf_col = "UF"

    mask = pd.Series(True, index=df.index)
    if status_col in df.columns:
        mask &= df[status_col].astype(str).str.upper().str.strip() == "FATURADO"
    if uf_col in df.columns:
        mask &= df[uf_col].astype(str).str.upper().str.strip() == "BA"

    df_ba = df[mask].copy()

    if df_ba.empty:
        result["ok"] = True
        result["erros"].append("Nenhum processo FATURADO com UF=BA encontrado no mês.")
        return result

    # ── Aba 1: Processos BA ──────────────────────────────────────────
    def _safe_col(df: pd.DataFrame, col: str) -> pd.Series:
        return df[col] if col in df.columns else pd.Series("", index=df.index)

    grp_cols = {
        "Processo": _safe_col(df_ba, "Processo"),
        "NF": _safe_col(df_ba, "Numero NF"),
        "Nome Cliente": _safe_col(df_ba, "Nome Cliente"),
        "Cidade": _safe_col(df_ba, "Cidade"),
        "UF": _safe_col(df_ba, "UF"),
        "Valor Realizado": pd.to_numeric(_safe_col(df_ba, "Valor Realizado"), errors="coerce").fillna(0),
        "Linha": _safe_col(df_ba, "Linha"),
    }

    df_grp_input = pd.DataFrame(grp_cols)

    df_processos = (
        df_grp_input.groupby("Processo", as_index=False)
        .agg(
            NF=("NF", lambda x: ", ".join(sorted(set(str(v) for v in x if str(v).strip())))),
            Nome_Cliente=("Nome Cliente", "first"),
            Cidade=("Cidade", "first"),
            UF=("UF", "first"),
            Valor_Realizado_Total=("Valor Realizado", "sum"),
            Linhas_de_Produto=("Linha", lambda x: ", ".join(sorted(set(str(v) for v in x if str(v).strip())))),
        )
        .rename(columns={
            "Nome_Cliente": "Nome Cliente",
            "Valor_Realizado_Total": "Valor Realizado Total",
            "Linhas_de_Produto": "Linhas de Produto",
        })
    )
    df_processos["Validado"] = ""

    # ── Aba 2: Detalhe Itens ─────────────────────────────────────────
    detalhe_cols = [
        "Processo", "Numero NF", "Código Produto", "Descrição Produto",
        "Linha", "Grupo", "Valor Realizado", "Cidade", "UF",
    ]
    available = [c for c in detalhe_cols if c in df_ba.columns]
    df_detalhe = df_ba[available].copy()

    # ── Gerar Excel ──────────────────────────────────────────────────
    if output_dir is None:
        root = Path(__file__).resolve().parent.parent
        output_dir = root / "saida" / f"{mes:02d}_{ano}"

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    arquivo = output_path / f"validacao_ba_{mes:02d}_{ano}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    _HEADER_FONT = Font(color="FFFFFF", bold=True)
    _VALIDADO_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    def _write_sheet(wb, title: str, df: pd.DataFrame, highlight_col: Optional[str] = None):
        ws = wb.create_sheet(title)
        cols = list(df.columns)
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for col_idx, col_name in enumerate(cols, 1):
                val = row[col_name]
                ws.cell(row=row_idx, column=col_idx, value=val if not pd.isna(val) else "")
                if highlight_col and col_name == highlight_col:
                    ws.cell(row=row_idx, column=col_idx).fill = _VALIDADO_FILL

        for col_idx in range(1, len(cols) + 1):
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    _write_sheet(wb, "Processos BA", df_processos, highlight_col="Validado")
    _write_sheet(wb, "Detalhe Itens", df_detalhe)

    wb.save(str(arquivo))

    result["ok"] = True
    result["arquivo"] = str(arquivo)
    result["num_processos"] = len(df_processos)
    result["num_itens"] = len(df_detalhe)
    return result
