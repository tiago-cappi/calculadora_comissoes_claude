"""
=============================================================================
Raiolanda — Cálculo de Comissão após Validação Manual
=============================================================================
Lê o Excel validado (validacao_ba_MM_AAAA.xlsx com coluna "Validado"),
filtra processos onde Validado == "Sim" e calcula comissão fixa por item.

Fórmula: Comissão = Valor Realizado × taxa_fixa_pct / 100
         Sem FC — taxa fixa direta.

Saída: saida/MM_AAAA/comissao_raiolanda_MM_AAAA.xlsx

Uso como script standalone:
    python scripts/raiolanda_comissao.py --mes 10 --ano 2025 \
        --validated-file saida/10_2025/validacao_ba_10_2025.xlsx
=============================================================================
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Optional

import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))


def _get_taxa_fixa(taxa_fixa_pct_override: Optional[float] = None) -> float:
    """Retorna a taxa fixa de Raiolanda do params.json ou o override fornecido."""
    if taxa_fixa_pct_override is not None:
        return float(taxa_fixa_pct_override)

    try:
        import scripts.config_manager as cm
        params = cm.get_params()
        val = params.get("raiolanda_taxa_fixa_pct")
        if val is not None:
            return float(val)
    except Exception:
        pass

    raise ValueError(
        "Taxa fixa de Raiolanda não configurada. "
        "Defina 'raiolanda_taxa_fixa_pct' em params ou passe --taxa-fixa."
    )


def execute(
    df_analise_comercial: pd.DataFrame,
    validated_file_path: str,
    mes: int,
    ano: int,
    taxa_fixa_pct: Optional[float] = None,
    output_dir: Optional[str] = None,
) -> dict:
    """Calcula comissão de Raiolanda a partir do Excel validado.

    Args:
        df_analise_comercial: DataFrame da AC já filtrado pelo mês/ano.
        validated_file_path: Caminho para o Excel validado pela analista.
        mes: Mês de apuração.
        ano: Ano de apuração.
        taxa_fixa_pct: Override da taxa fixa (%). Se None, usa params.json.
        output_dir: Diretório de saída. Se None, usa saida/MM_AAAA/.

    Returns:
        dict com 'ok', 'arquivo', 'total_comissao', 'num_processos', 'num_itens', 'erros'.
    """
    result: dict = {
        "ok": False, "arquivo": None,
        "total_comissao": 0.0, "num_processos": 0, "num_itens": 0, "erros": [],
    }

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        result["erros"].append(f"openpyxl não disponível: {exc}")
        return result

    # ── Carregar taxa fixa ───────────────────────────────────────────
    try:
        taxa = _get_taxa_fixa(taxa_fixa_pct)
    except ValueError as exc:
        result["erros"].append(str(exc))
        return result

    # ── Ler Excel validado ───────────────────────────────────────────
    validated_path = Path(validated_file_path)
    if not validated_path.exists():
        result["erros"].append(f"Arquivo validado não encontrado: {validated_file_path}")
        return result

    try:
        df_validado = pd.read_excel(str(validated_path), sheet_name="Processos BA", dtype=str)
    except Exception as exc:
        result["erros"].append(f"Erro ao ler aba 'Processos BA': {exc}")
        return result

    if "Validado" not in df_validado.columns or "Processo" not in df_validado.columns:
        result["erros"].append("Colunas 'Validado' e/ou 'Processo' não encontradas no Excel validado.")
        return result

    processos_validados = set(
        df_validado.loc[
            df_validado["Validado"].astype(str).str.strip().str.lower() == "sim",
            "Processo",
        ]
        .astype(str).str.strip().unique()
    )

    if not processos_validados:
        result["erros"].append("Nenhum processo com Validado='Sim' encontrado.")
        return result

    # ── Filtrar AC pelos processos validados ─────────────────────────
    if df_analise_comercial is None or df_analise_comercial.empty:
        result["erros"].append("Análise Comercial vazia.")
        return result

    df_ac = df_analise_comercial.copy()

    if "Processo" not in df_ac.columns:
        result["erros"].append("Coluna 'Processo' não encontrada na Análise Comercial.")
        return result

    df_itens = df_ac[
        df_ac["Processo"].astype(str).str.strip().isin(processos_validados) &
        (df_ac.get("Status Processo", pd.Series("", index=df_ac.index))
         .astype(str).str.upper().str.strip() == "FATURADO")
    ].copy()

    if df_itens.empty:
        result["erros"].append("Nenhum item FATURADO encontrado para os processos validados.")
        return result

    # ── Calcular comissão ────────────────────────────────────────────
    def _safe_float(val) -> float:
        try:
            return float(str(val).replace(",", ".").strip())
        except (ValueError, TypeError):
            return 0.0

    df_itens["_valor_real"] = df_itens["Valor Realizado"].apply(_safe_float) if "Valor Realizado" in df_itens.columns else 0.0
    df_itens["_comissao"] = df_itens["_valor_real"] * taxa / 100.0

    total_comissao = df_itens["_comissao"].sum()

    # ── Montar DataFrames de saída ───────────────────────────────────
    resumo_cols = ["Processo", "Nome Cliente", "Cidade", "UF"]
    resumo_available = [c for c in resumo_cols if c in df_itens.columns]
    df_resumo = (
        df_itens.groupby("Processo", as_index=False)
        .agg(
            **{c: (c, "first") for c in resumo_available if c != "Processo"},
            Valor_Total=("_valor_real", "sum"),
            Comissao_Total=("_comissao", "sum"),
            Num_Itens=("_comissao", "count"),
        )
        .rename(columns={"Valor_Total": "Valor Total (R$)", "Comissao_Total": "Comissão (R$)", "Num_Itens": "Nº Itens"})
    )
    df_resumo["Taxa Fixa (%)"] = taxa
    total_row = pd.DataFrame([{
        "Processo": "TOTAL",
        "Valor Total (R$)": df_resumo["Valor Total (R$)"].sum(),
        "Comissão (R$)": total_comissao,
        "Taxa Fixa (%)": taxa,
    }])
    df_resumo = pd.concat([df_resumo, total_row], ignore_index=True)

    detalhe_cols = ["Processo", "Numero NF", "Código Produto", "Descrição Produto",
                    "Linha", "Grupo", "Cidade", "UF"]
    det_available = [c for c in detalhe_cols if c in df_itens.columns]
    df_detalhe = df_itens[det_available + ["_valor_real", "_comissao"]].rename(columns={
        "_valor_real": "Valor Realizado (R$)",
        "_comissao": "Comissão (R$)",
    }).copy()
    df_detalhe["Taxa Fixa (%)"] = taxa
    df_detalhe["FC"] = "N/A (taxa fixa)"

    # ── Gerar Excel ──────────────────────────────────────────────────
    if output_dir is None:
        output_dir = ROOT / "saida" / f"{mes:02d}_{ano}"

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    arquivo = output_path / f"comissao_raiolanda_{mes:02d}_{ano}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    _HEADER_FONT = Font(color="FFFFFF", bold=True)
    _TOTAL_FONT = Font(bold=True)
    _MONEY_FMT = '#,##0.00'
    _PCT_FMT = '0.00'

    def _write_sheet(wb, title: str, df: pd.DataFrame, money_cols: list, pct_cols: list):
        ws = wb.create_sheet(title)
        cols = list(df.columns)
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            is_total = str(row.get("Processo", "")).upper() == "TOTAL"
            for col_idx, col_name in enumerate(cols, 1):
                val = row[col_name]
                try:
                    cell_val = float(val) if col_name in money_cols + pct_cols else val
                except (ValueError, TypeError):
                    cell_val = val
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_val if not (pd.isna(cell_val) if not isinstance(cell_val, str) else False) else "")
                if col_name in money_cols:
                    cell.number_format = _MONEY_FMT
                elif col_name in pct_cols:
                    cell.number_format = _PCT_FMT
                if is_total:
                    cell.font = _TOTAL_FONT

        for col_idx in range(1, len(cols) + 1):
            max_len = max(
                (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 50)

    _write_sheet(
        wb, "Resumo por Processo", df_resumo,
        money_cols=["Valor Total (R$)", "Comissão (R$)"],
        pct_cols=["Taxa Fixa (%)"],
    )
    _write_sheet(
        wb, "Detalhe por Item", df_detalhe,
        money_cols=["Valor Realizado (R$)", "Comissão (R$)"],
        pct_cols=["Taxa Fixa (%)"],
    )

    wb.save(str(arquivo))

    result["ok"] = True
    result["arquivo"] = str(arquivo)
    result["total_comissao"] = round(total_comissao, 2)
    result["num_processos"] = len(processos_validados)
    result["num_itens"] = len(df_itens)
    return result


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Calcula comissão de Raiolanda (BA) após validação manual.")
    parser.add_argument("--mes", type=int, required=True)
    parser.add_argument("--ano", type=int, required=True)
    parser.add_argument("--validated-file", required=True, help="Caminho do Excel validado pela analista.")
    parser.add_argument("--taxa-fixa", type=float, default=None, help="Override da taxa fixa em %.")
    args = parser.parse_args()

    import scripts.loaders as loader

    dados_dir = ROOT / "dados_entrada"
    ac_bytes = (dados_dir / "analise-comercial.xlsx").read_bytes()
    cp_bytes = (dados_dir / "Classificação de Produtos.xlsx").read_bytes()

    loader_result = loader.execute(
        mes=args.mes,
        ano=args.ano,
        file_analise_comercial=ac_bytes,
        file_classificacao_produtos=cp_bytes,
    )

    res = execute(
        df_analise_comercial=loader_result.analise_comercial,
        validated_file_path=args.validated_file,
        mes=args.mes,
        ano=args.ano,
        taxa_fixa_pct=args.taxa_fixa,
    )

    if res["ok"]:
        print(f"✓ Comissão Raiolanda gerada: {res['arquivo']}")
        print(f"  Processos: {res['num_processos']} | Itens: {res['num_itens']} | Total: R$ {res['total_comissao']:,.2f}")
    else:
        print("✗ Erros:")
        for e in res["erros"]:
            print(f"  - {e}")
        sys.exit(1)
