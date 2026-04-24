"""
=============================================================================
SKILL: Robô de Comissões — Script: Exportação Excel por Colaborador
=============================================================================
Módulo   : excel_export
Versão   : 1.0.0

Descrição
---------
Gera **um arquivo Excel por colaborador** com 3 abas formatadas como
tabelas Excel (ListObject) com auto-filtros e totais, para visualização
e auditoria das comissões calculadas.

Abas:
  1. Resumo — consolidado por linha de negócio
  2. Detalhe por Item — cada item comissionado
  3. FC Componentes — componentes do Fator de Correção

Dependências
------------
- openpyxl
- ComissaoFaturamentoResult (scripts.comissao_faturamento)
- FCResultSet               (scripts.fc_calculator)
=============================================================================
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Lazy imports to avoid circular deps — resolved inside execute()
# from scripts.comissao_faturamento import ComissaoFaturamentoResult
# from scripts.fc_calculator import FCResultSet, FCResult


# ═══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _sanitize_filename(name: str) -> str:
    """Remove acentos, caracteres especiais e espaços para nome de arquivo."""
    # NFD → remove combining marks
    nfkd = unicodedata.normalize("NFKD", name)
    ascii_name = nfkd.encode("ascii", "ignore").decode("ascii")
    # Replace spaces/special chars with underscore
    clean = re.sub(r"[^\w]", "_", ascii_name)
    # Collapse multiple underscores
    clean = re.sub(r"_+", "_", clean).strip("_")
    return clean


def _fmt_money(value: Any) -> float:
    """Garante float para colunas monetárias."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _fmt_pct(value: Any) -> float:
    """Garante float para colunas de percentual (já em %)."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _taxa_maxima_comissao_potencial_pct(comissao: Dict[str, Any]) -> float:
    """Taxa máxima potencial = taxa de rateio x fatia do cargo."""
    taxa_rateio_pct = _fmt_pct(comissao.get("taxa_rateio_pct", 0))
    fatia_cargo_pct = _fmt_pct(comissao.get("fatia_cargo_pct", 0))
    return taxa_rateio_pct * fatia_cargo_pct / 100.0


def _taxa_efetiva_comissao_pct(comissao: Dict[str, Any]) -> float:
    """Taxa efetiva = taxa máxima potencial x FC aplicado."""
    fc_aplicado = _fmt_money(comissao.get("fc_final", 0))
    return _taxa_maxima_comissao_potencial_pct(comissao) * fc_aplicado


def _media_ponderada_taxa_pct(comissoes: List[Dict[str, Any]], value_getter) -> float:
    """Calcula média ponderada por valor_item para agregações do Resumo."""
    soma_pesos = 0.0
    soma_ponderada = 0.0
    for comissao in comissoes:
        peso = abs(_fmt_money(comissao.get("valor_item", 0)))
        taxa_pct = _fmt_pct(value_getter(comissao))
        soma_pesos += peso
        soma_ponderada += peso * taxa_pct
    return soma_ponderada / soma_pesos if soma_pesos > 0 else 0.0


_MONEY_FMT = '#,##0.00'
_PCT_FMT = '0.00"%"'
_FACTOR_FMT = '0.0000'

_HEADER_FONT = Font(name="Calibri", bold=True, size=14)
_SUBHEADER_FONT = Font(name="Calibri", bold=False, size=11, italic=True, color="666666")
_TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)


def _auto_width(ws, start_row: int, end_row: int, num_cols: int, min_width: int = 10):
    """Ajusta largura das colunas baseado no conteúdo."""
    for col_idx in range(1, num_cols + 1):
        max_len = min_width
        for row_idx in range(start_row, end_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        # Cap column width
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 35)


def _add_table(ws, start_row: int, end_row: int, num_cols: int, table_name: str):
    """Adiciona uma tabela Excel (ListObject) com estilo e auto-filtro."""
    ref = f"A{start_row}:{get_column_letter(num_cols)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = _TABLE_STYLE
    ws.add_table(table)


def _apply_format(ws, col_idx: int, start_row: int, end_row: int, fmt: str):
    """Aplica formato numérico a uma coluna."""
    for row_idx in range(start_row, end_row + 1):
        ws.cell(row=row_idx, column=col_idx).number_format = fmt


# ═══════════════════════════════════════════════════════════════════════════════
# DATA CLASS
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class ExcelExportResult:
    """Resultado da exportação Excel."""
    arquivos_gerados: List[str] = field(default_factory=list)
    diretorio: str = ""
    warnings: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return len(self.errors) == 0

    def summary(self) -> str:
        lines = [
            f"{'='*65}",
            f"  EXPORTAÇÃO EXCEL — Resumo",
            f"{'='*65}",
            f"  Diretório: {self.diretorio}",
            f"  Arquivos gerados: {len(self.arquivos_gerados)}",
        ]
        for f in self.arquivos_gerados:
            lines.append(f"    ✓ {f}")
        if self.warnings:
            lines.append(f"\n  ⚠ Avisos ({len(self.warnings)}):")
            for w in self.warnings[:10]:
                lines.append(f"    • {w}")
        if self.errors:
            lines.append(f"\n  ✖ Erros ({len(self.errors)}):")
            for e in self.errors[:10]:
                lines.append(f"    • {e}")
        lines.append(f"{'='*65}")
        return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET BUILDERS
# ═══════════════════════════════════════════════════════════════════════════════

def _build_fc_groups(comissoes: List[Dict], fc_results: List) -> Dict[str, Dict]:
    """Constrói grupos FC agrupando FCResults com composição idêntica.

    Dois FCResults estão no mesmo grupo se possuem o mesmo fc_rampa,
    o que implica que resolveram para as mesmas metas/realizados.

    Returns:
        Dict[label, {"fc": FCResult, "hierarquia_key_set": set, "comissoes": list}]
        Labels são letras: "A", "B", "C", ...
    """
    from collections import OrderedDict

    # Agrupar comissões por hierarquia_key → fc_rampa
    hk_to_fc: Dict[str, Any] = {}
    for fcr in fc_results:
        hk = getattr(fcr, "hierarquia_key", "") or fcr.linha
        hk_to_fc[hk] = fcr

    # Agrupar FCResults por composição (fc_rampa como proxy)
    fc_groups_raw: Dict[float, Dict] = OrderedDict()
    for fcr in sorted(fc_results, key=lambda x: getattr(x, "hierarquia_key", "") or x.linha):
        key = round(fcr.fc_rampa, 8)
        if key not in fc_groups_raw:
            fc_groups_raw[key] = {
                "fc": fcr,
                "hierarquia_key_set": set(),
                "comissoes": [],
            }
        hk = getattr(fcr, "hierarquia_key", "") or fcr.linha
        fc_groups_raw[key]["hierarquia_key_set"].add(hk)

    # Atribuir comissões a grupos
    for c in comissoes:
        hk = c.get("hierarquia_key", "") or c.get("linha", "")
        fcr = hk_to_fc.get(hk)
        if fcr:
            key = round(fcr.fc_rampa, 8)
            if key in fc_groups_raw:
                fc_groups_raw[key]["comissoes"].append(c)
                continue
        # Fallback: comissão sem FC match (improvável)
        if fc_groups_raw:
            first_key = next(iter(fc_groups_raw))
            fc_groups_raw[first_key]["comissoes"].append(c)

    # Nomear grupos: A, B, C, ...
    labels = [chr(65 + i) for i in range(26)]
    groups = OrderedDict()
    for i, (k, v) in enumerate(fc_groups_raw.items()):
        label = labels[i] if i < 26 else f"G{i+1}"
        groups[label] = v

    return groups


def _fc_group_label_for_hk(hk: str, fc_groups: Dict[str, Dict]) -> str:
    """Retorna a label do grupo FC ("A", "B", ...) para uma hierarquia_key."""
    for label, grp in fc_groups.items():
        if hk in grp["hierarquia_key_set"]:
            return label
    return ""


def _shortest_common_prefix(keys: set) -> str:
    """Encontra o prefixo comum mais curto de um conjunto de hierarquia_keys."""
    if not keys:
        return ""
    parts_list = [k.split("/") for k in keys]
    min_len = min(len(p) for p in parts_list)
    common = []
    for i in range(min_len):
        vals = set(p[i] for p in parts_list)
        if len(vals) == 1:
            common.append(vals.pop())
        else:
            break
    return "/".join(common) if common else next(iter(keys)).split("/")[0]


def _build_resumo(wb: openpyxl.Workbook, nome: str, cargo: str,
                  comissoes: List[Dict], fc_results: List, mes: int, ano: int,
                  fc_groups: Dict[str, Dict]):
    """Aba 1 — Resumo consolidado por Grupo FC."""
    ws = wb.active
    ws.title = "Resumo"

    # Header
    ws.cell(row=1, column=1, value=f"{nome}").font = _HEADER_FONT
    ws.cell(row=2, column=1, value=f"{cargo} — Comissões {mes:02d}/{ano}").font = _SUBHEADER_FONT

    # Tabela começa na linha 4
    headers = [
        "Grupo FC", "Hierarquia", "Qtd Itens",
        "Comissão Potencial (R$)", "Taxa Máx. Comissão Potencial (%)",
        "FC Modo", "FC Rampa", "FC Final", "Taxa Efetiva Comissão (%)",
        "Comissão Final (R$)",
    ]
    start_row = 4
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=h)

    row = start_row + 1
    for label in sorted(fc_groups.keys()):
        grp = fc_groups[label]
        fcr = grp["fc"]
        grp_comissoes = grp["comissoes"]
        hierarquia_label = _shortest_common_prefix(grp["hierarquia_key_set"])

        pot = sum(_fmt_money(c.get("comissao_potencial", 0)) for c in grp_comissoes)
        final = sum(_fmt_money(c.get("comissao_final", 0)) for c in grp_comissoes)
        taxa_pot_pct = _media_ponderada_taxa_pct(grp_comissoes, _taxa_maxima_comissao_potencial_pct)
        taxa_efetiva_pct = _media_ponderada_taxa_pct(grp_comissoes, _taxa_efetiva_comissao_pct)

        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=hierarquia_label)
        ws.cell(row=row, column=3, value=len(grp_comissoes))
        ws.cell(row=row, column=4, value=pot)
        ws.cell(row=row, column=5, value=taxa_pot_pct)
        ws.cell(row=row, column=6, value=fcr.modo)
        ws.cell(row=row, column=7, value=fcr.fc_rampa)
        ws.cell(row=row, column=8, value=fcr.fc_final)
        ws.cell(row=row, column=9, value=taxa_efetiva_pct)
        ws.cell(row=row, column=10, value=final)
        row += 1

    # Linha de TOTAL
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=3, value=len(comissoes))
    ws.cell(row=row, column=3).font = Font(bold=True)
    total_pot = sum(_fmt_money(c.get("comissao_potencial", 0)) for c in comissoes)
    total_final = sum(_fmt_money(c.get("comissao_final", 0)) for c in comissoes)
    total_taxa_pot_pct = _media_ponderada_taxa_pct(comissoes, _taxa_maxima_comissao_potencial_pct)
    total_taxa_efetiva_pct = _media_ponderada_taxa_pct(comissoes, _taxa_efetiva_comissao_pct)
    ws.cell(row=row, column=4, value=total_pot)
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=5, value=total_taxa_pot_pct)
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=9, value=total_taxa_efetiva_pct)
    ws.cell(row=row, column=9).font = Font(bold=True)
    ws.cell(row=row, column=10, value=total_final)
    ws.cell(row=row, column=10).font = Font(bold=True)

    end_row = row

    # Formatos
    _apply_format(ws, 4, start_row + 1, end_row, _MONEY_FMT)   # Potencial
    _apply_format(ws, 5, start_row + 1, end_row, _PCT_FMT)      # Taxa Potencial
    _apply_format(ws, 7, start_row + 1, end_row, _FACTOR_FMT)   # FC Rampa
    _apply_format(ws, 8, start_row + 1, end_row, _FACTOR_FMT)   # FC Final
    _apply_format(ws, 9, start_row + 1, end_row, _PCT_FMT)      # Taxa Efetiva
    _apply_format(ws, 10, start_row + 1, end_row, _MONEY_FMT)   # Final

    # Tabela Excel (sem a linha de TOTAL)
    if end_row > start_row + 1:
        _add_table(ws, start_row, end_row - 1, len(headers), "TblResumo")

    _auto_width(ws, start_row, end_row, len(headers))


def _build_detalhe(wb: openpyxl.Workbook, comissoes: List[Dict],
                   fc_groups: Dict[str, Dict]):
    """Aba 2 — Detalhe por Item."""
    ws = wb.create_sheet("Detalhe por Item")

    headers = [
        "Grupo FC", "Processo", "NF", "Cód. Produto", "Descrição Produto",
        "Linha", "Grupo", "Subgrupo",
        "Tipo Mercadoria", "Fabricante", "Aplicação", "Cliente",
        "Valor Item (R$)", "Taxa Rateio (%)", "Fatia Cargo (%)",
        "Taxa Máx. Comissão Potencial (%)", "FC Aplicado",
        "Taxa Efetiva Comissão (%)", "Comissão Potencial (R$)",
        "Comissão Final (R$)", "Observação",
    ]

    start_row = 1
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col_idx, value=h)

    for i, c in enumerate(comissoes, start=start_row + 1):
        # Injetar grupo FC
        hk = c.get("hierarquia_key", "") or c.get("linha", "")
        c["_grupo_fc"] = _fc_group_label_for_hk(hk, fc_groups)
        row_values = [
            c.get("_grupo_fc", ""), c.get("processo", ""), c.get("numero_nf", ""), c.get("codigo_produto", ""), c.get("descricao_produto", ""),
            c.get("linha", ""), c.get("grupo", ""), c.get("subgrupo", ""),
            c.get("tipo_mercadoria", ""), c.get("fabricante", ""), c.get("aplicacao", ""), c.get("nome_cliente", ""),
            c.get("valor_item", ""), c.get("taxa_rateio_pct", ""), c.get("fatia_cargo_pct", ""),
            _taxa_maxima_comissao_potencial_pct(c), c.get("fc_final", ""),
            _taxa_efetiva_comissao_pct(c), c.get("comissao_potencial", ""),
            c.get("comissao_final", ""), c.get("observacao", ""),
        ]
        for col_idx, val in enumerate(row_values, 1):
            ws.cell(row=i, column=col_idx, value=val)

    end_row = start_row + len(comissoes)

    # Colunas monetárias: 13 (Valor Item), 19 (Pot), 20 (Final)
    for col_idx in [13, 19, 20]:
        _apply_format(ws, col_idx, start_row + 1, end_row, _MONEY_FMT)
    # Colunas pct: 14 (Taxa), 15 (Fatia), 16 (Taxa Potencial), 18 (Taxa Efetiva)
    for col_idx in [14, 15, 16, 18]:
        _apply_format(ws, col_idx, start_row + 1, end_row, _PCT_FMT)
    # Fator: 17 (FC)
    _apply_format(ws, 17, start_row + 1, end_row, _FACTOR_FMT)

    if end_row > start_row:
        _add_table(ws, start_row, end_row, len(headers), "TblDetalhe")

    _auto_width(ws, start_row, end_row, len(headers))

    _PDIR_FILL = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    for i, c in enumerate(comissoes, start=start_row + 1):
        if c.get("is_pdir", False):
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=i, column=col_idx).fill = _PDIR_FILL
            obs_col = headers.index("Observação") + 1
            obs_cell = ws.cell(row=i, column=obs_col)
            obs_cell.value = f"PDIR — {obs_cell.value}" if obs_cell.value else "PDIR"


# ═══════════════════════════════════════════════════════════════════════════════
# AUDITORIA — Nova aba: Itens AC (Auditoria)
# ═══════════════════════════════════════════════════════════════════════════════

_AUDIT_AC_COLS: List[str] = [
    "Código Produto", "Descrição Produto", "Qtde Solicitada", "Qtde Atendida",
    "Preço Unitário", "Operação", "Processo", "Numero NF", "Status Processo",
    "Dt Entrada", "Status da NF", "Data Aceite", "Dt Aprovação", "Dt Emissão",
    "Valor Orçado", "Valor Realizado", "Consultor Interno", "Representante-pedido",
    "Gerente Comercial-Pedido", "Aplicação Mat./Serv.", "Cliente", "Nome Cliente",
    "Cidade", "UF",
]

_AUDIT_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_AUDIT_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _audit_norm(v: Any) -> str:
    if v is None:
        return ""
    try:
        import pandas as _pd
        if _pd.isna(v):
            return ""
    except Exception:
        pass
    return str(v).strip().upper()


def _audit_cell_value(v: Any) -> Any:
    """Converte valor para algo serializável pelo openpyxl."""
    if v is None:
        return ""
    try:
        import pandas as _pd
        if _pd.isna(v):
            return ""
        if hasattr(v, "to_pydatetime"):
            return v.to_pydatetime()
    except Exception:
        pass
    return v


def _build_auditoria_ac(wb: openpyxl.Workbook, comissoes: List[Dict], df_ac: Any):
    """Aba extra — Itens da AC do mês com colorização verde/vermelho por comissão.

    Verde: AC item (Processo, Numero NF, Código Produto) gerou comissão para
        este colaborador.
    Vermelho: não gerou.
    Ordenação: verdes primeiro, vermelhos depois; dentro de cada grupo por Processo.
    """
    ws = wb.create_sheet("Itens AC (Auditoria)")

    if df_ac is None or getattr(df_ac, "empty", True):
        ws.cell(row=1, column=1, value="AC indisponível — aba de auditoria vazia.")
        return

    chaves_comissao = {
        (
            _audit_norm(c.get("processo")),
            _audit_norm(c.get("numero_nf")),
            _audit_norm(c.get("codigo_produto")),
        )
        for c in comissoes
    }

    cols_disponiveis = [c for c in _AUDIT_AC_COLS if c in df_ac.columns]
    cols_faltantes = [c for c in _AUDIT_AC_COLS if c not in df_ac.columns]

    # Header
    for col_idx, h in enumerate(_AUDIT_AC_COLS, 1):
        ws.cell(row=1, column=col_idx, value=h).font = Font(bold=True)

    # Pré-computar cor e chave de ordenação para cada linha
    linhas: List[tuple] = []  # (is_green, processo_sort, row_dict)
    for _, row in df_ac.iterrows():
        proc = _audit_norm(row.get("Processo"))
        nf = _audit_norm(row.get("Numero NF"))
        cod = _audit_norm(row.get("Código Produto"))
        is_green = (proc, nf, cod) in chaves_comissao
        linhas.append((is_green, proc, row))

    # Ordenação: verde primeiro (is_green True → 0), depois por Processo
    linhas.sort(key=lambda t: (0 if t[0] else 1, t[1]))

    # Escrita
    start_data = 2
    for row_idx, (is_green, _, row) in enumerate(linhas, start=start_data):
        fill = _AUDIT_GREEN_FILL if is_green else _AUDIT_RED_FILL
        for col_idx, col_name in enumerate(_AUDIT_AC_COLS, 1):
            valor = row.get(col_name, "") if col_name in cols_disponiveis else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=_audit_cell_value(valor))
            cell.fill = fill
            if col_name in ("Valor Orçado", "Valor Realizado", "Preço Unitário"):
                cell.number_format = _MONEY_FMT

    end_row = start_data + len(linhas) - 1
    if end_row >= start_data:
        ws.freeze_panes = "A2"
        _auto_width(ws, 1, end_row, len(_AUDIT_AC_COLS))

    if cols_faltantes:
        # Nota ao lado da primeira linha avisando colunas ausentes no AC carregado
        nota_col = len(_AUDIT_AC_COLS) + 2
        ws.cell(row=1, column=nota_col,
                value=f"Colunas ausentes na AC: {', '.join(cols_faltantes)}").font = Font(
                    italic=True, color="888888")


def _build_calculo_fc(wb: openpyxl.Workbook, nome: str, fc_results: List,
                      fc_groups: Dict[str, Dict]):
    """Aba 3 — Cálculo FC por grupo (seções com componentes + escada)."""
    from scripts.fc_calculator import gerar_degraus_escada

    ws = wb.create_sheet("Cálculo FC")

    _BOLD = Font(bold=True)
    _SECTION_FONT = Font(name="Calibri", bold=True, size=12)
    _SECTION_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    _STEP_CURRENT_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    _RESULT_FONT = Font(bold=True, size=11)

    row = 1
    for label in sorted(fc_groups.keys()):
        grp = fc_groups[label]
        fcr = grp["fc"]
        grp_comissoes = grp["comissoes"]
        hierarquia_label = _shortest_common_prefix(grp["hierarquia_key_set"])
        n_itens = len(grp_comissoes)
        total_valor = sum(_fmt_money(c.get("valor_item", 0)) for c in grp_comissoes)

        # ── Section Header ──
        header_text = f"GRUPO {label} — {hierarquia_label}    ({n_itens} itens, R$ {total_valor:,.2f})"
        ws.cell(row=row, column=1, value=header_text).font = _SECTION_FONT
        for col in range(1, 9):
            ws.cell(row=row, column=col).fill = _SECTION_FILL
        row += 1

        # ── Componentes Table ──
        comp_headers = ["Componente", "Peso (%)", "Realizado", "Meta",
                        "Atingimento (%)", "Ating. c/ Cap (%)", "Contribuição"]
        for col_idx, h in enumerate(comp_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = _BOLD
        row += 1

        comps_ativos = [c for c in fcr.componentes if c.peso > 0]
        comp_start_row = row
        for comp in comps_ativos:
            ws.cell(row=row, column=1, value=comp.nome)
            ws.cell(row=row, column=2, value=comp.peso * 100)
            ws.cell(row=row, column=3, value=comp.realizado)
            ws.cell(row=row, column=4, value=comp.meta)
            ws.cell(row=row, column=5, value=comp.atingimento * 100)
            ws.cell(row=row, column=6, value=comp.atingimento_cap * 100)
            ws.cell(row=row, column=7, value=comp.contribuicao)
            row += 1
        comp_end_row = row - 1

        # Formatos dos componentes
        if comp_end_row >= comp_start_row:
            _apply_format(ws, 2, comp_start_row, comp_end_row, '0.00')      # Peso
            _apply_format(ws, 3, comp_start_row, comp_end_row, _MONEY_FMT)  # Realizado
            _apply_format(ws, 4, comp_start_row, comp_end_row, _MONEY_FMT)  # Meta
            _apply_format(ws, 5, comp_start_row, comp_end_row, '0.00')      # Ating
            _apply_format(ws, 6, comp_start_row, comp_end_row, '0.00')      # Ating Cap
            _apply_format(ws, 7, comp_start_row, comp_end_row, _FACTOR_FMT) # Contribuição

        # ── FC Rampa ──
        ws.cell(row=row, column=1, value="FC Rampa =")
        ws.cell(row=row, column=1).font = _RESULT_FONT
        ws.cell(row=row, column=2, value=fcr.fc_rampa)
        ws.cell(row=row, column=2).number_format = _FACTOR_FMT
        ws.cell(row=row, column=2).font = _RESULT_FONT
        row += 1

        # ── Escada (se aplicável) ──
        if fcr.modo == "ESCADA" and fcr.escada_num_degraus:
            n_deg = fcr.escada_num_degraus
            piso_val = fcr.escada_piso or 0.0
            degrau_atual = fcr.escada_degrau_indice

            row += 1  # Espaço
            ws.cell(row=row, column=1,
                    value=f"ESCADA ({n_deg} degraus, piso {piso_val:.0%})")
            ws.cell(row=row, column=1).font = _BOLD
            row += 1

            # Headers da escada
            for col_idx, h in enumerate(["Degrau", "Valor", ""], 1):
                ws.cell(row=row, column=col_idx, value=h).font = _BOLD
            row += 1

            steps = gerar_degraus_escada(n_deg, piso_val)
            for k, step_val in enumerate(steps):
                ws.cell(row=row, column=1, value=k)
                ws.cell(row=row, column=2, value=f"{step_val:.4f} ({step_val:.0%})")
                if k == degrau_atual:
                    ws.cell(row=row, column=3, value="← ATUAL")
                    ws.cell(row=row, column=3).font = _BOLD
                    for col in range(1, 4):
                        ws.cell(row=row, column=col).fill = _STEP_CURRENT_FILL
                row += 1

        # ── FC Final ──
        row += 1
        ws.cell(row=row, column=1, value="FC Final =")
        ws.cell(row=row, column=1).font = _RESULT_FONT
        ws.cell(row=row, column=2, value=fcr.fc_final)
        ws.cell(row=row, column=2).number_format = _FACTOR_FMT
        ws.cell(row=row, column=2).font = _RESULT_FONT
        ws.cell(row=row, column=3, value=f"({fcr.modo})")
        row += 2  # Espaço entre seções

    # Ajustar larguras
    if row > 1:
        _auto_width(ws, 1, row - 1, 8)


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN EXECUTE
# ═══════════════════════════════════════════════════════════════════════════════

def execute(
    result_fat,   # ComissaoFaturamentoResult
    result_fc,    # FCResultSet
    mes: int,
    ano: int,
    output_dir: Optional[str] = None,
    df_ac: Any = None,
) -> ExcelExportResult:
    """Gera um arquivo Excel por colaborador comissionado.

    Args:
        result_fat: Resultado do cálculo de comissão por faturamento.
        result_fc: Resultado do cálculo de FC.
        mes: Mês de apuração.
        ano: Ano de apuração.
        output_dir: Diretório de saída (padrão: ./saida/MM_AAAA/).

    Returns:
        ExcelExportResult com lista de arquivos gerados.
    """
    result = ExcelExportResult()

    # Diretório de saída
    if output_dir:
        out_path = Path(output_dir)
    else:
        out_path = Path(".") / "saida" / f"{mes:02d}_{ano}"

    try:
        out_path.mkdir(parents=True, exist_ok=True)
    except PermissionError as exc:
        result.errors.append(
            f"PermissionError ao criar diretório '{out_path}': {exc}. "
            f"Verifique se a pasta está aberta no Explorer ou se outro processo a está usando."
        )
        result.diretorio = str(out_path)
        return result
    except OSError as exc:
        result.errors.append(f"OSError ao criar diretório '{out_path}': {exc}")
        result.diretorio = str(out_path)
        return result

    result.diretorio = str(out_path)

    # Agrupar comissões por colaborador
    by_colab: Dict[str, List[Dict]] = {}
    for c in result_fat.comissoes:
        nome = c.get("nome", "")
        if nome:
            by_colab.setdefault(nome, []).append(c)

    if not by_colab:
        result.warnings.append("Nenhuma comissão para exportar.")
        return result

    for nome in sorted(by_colab.keys()):
        comissoes = by_colab[nome]
        cargo = comissoes[0].get("cargo", "N/A")

        # Encontrar FC results deste colaborador
        fc_results = [
            r for r in result_fc.resultados
            if r.colaborador == nome
        ]

        # Construir grupos FC
        fc_groups = _build_fc_groups(comissoes, fc_results)

        # Criar workbook
        wb = openpyxl.Workbook()

        try:
            _build_resumo(wb, nome, cargo, comissoes, fc_results, mes, ano, fc_groups)
            _build_detalhe(wb, comissoes, fc_groups)
            _build_calculo_fc(wb, nome, fc_results, fc_groups)
            if df_ac is not None:
                _build_auditoria_ac(wb, comissoes, df_ac)

            # Salvar
            filename = f"comissao_{_sanitize_filename(nome)}_{mes:02d}_{ano}.xlsx"
            filepath = out_path / filename
            wb.save(str(filepath))
            result.arquivos_gerados.append(filename)

        except Exception as e:
            result.errors.append(f"Erro ao gerar Excel para {nome}: {e}")
        finally:
            wb.close()

    return result
