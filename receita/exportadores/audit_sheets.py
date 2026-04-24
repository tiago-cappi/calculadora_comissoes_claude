"""
receita/exportadores/audit_sheets.py — Abas de auditoria Supabase.

Gera 3 abas espelhando as tabelas de historico usadas pelo pipeline:
    - historico_comissoes              -> "Aud. Historico Comissoes"
    - historico_processo_pai           -> "Aud. Processos Pai"
    - historico_pagamentos_processo_pai -> "Aud. Pagamentos Pai"

As funcoes recebem dataclasses ja hidratadas (nenhuma chamada Supabase aqui).
Campos JSON grandes sao truncados na celula e o valor completo vai para
o Comment, preservando auditabilidade sem estourar a largura de coluna.
"""

from __future__ import annotations

from datetime import datetime
from typing import Any, Iterable, List, Optional

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from receita.schemas.historico import (
    HistoricoComissao,
    HistoricoPagamentoProcessoPai,
    HistoricoProcessoPai,
)


_MONEY_FMT = "#,##0.00"
_FACTOR_FMT = "0.0000"
_DATE_FMT = "dd/mm/yyyy"
_DATETIME_FMT = "dd/mm/yyyy hh:mm"

_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
_HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_ITALIC_FONT = Font(name="Calibri", italic=True, color="888888")

_JSON_MAX_CELL = 200
_COMMENT_MAX = 30000  # limite pratico do openpyxl Comment


def _write_header(ws, row: int, headers: List[str]) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _auto_width(ws, end_row: int, num_cols: int, max_width: int = 38) -> None:
    for col_idx in range(1, num_cols + 1):
        max_len = 10
        for row_idx in range(1, end_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


def _add_table(ws, start_row: int, end_row: int, num_cols: int, table_name: str) -> None:
    ref = f"A{start_row}:{get_column_letter(num_cols)}{end_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)


def _render_json_cell(ws, row: int, col: int, raw_value: Any) -> None:
    """Escreve JSON truncado na celula com o valor completo em Comment."""
    text = "" if raw_value is None else str(raw_value)
    if not text or text in ("[]", "{}"):
        ws.cell(row=row, column=col, value="")
        return

    truncado = text[:_JSON_MAX_CELL]
    if len(text) > _JSON_MAX_CELL:
        truncado += "..."
    cell = ws.cell(row=row, column=col, value=truncado)

    comment_text = text if len(text) <= _COMMENT_MAX else text[:_COMMENT_MAX] + "\n...[truncado]"
    try:
        cell.comment = Comment(comment_text, "auditoria")
    except Exception:
        pass


def _bool_label(value: Optional[bool]) -> str:
    if value is None:
        return "—"
    return "Sim" if value else "Nao"


def _fmt_datetime(value: Optional[datetime]) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    return value


def _write_unavailable_note(ws, titulo: str, motivo: str) -> None:
    """Quando dados do Supabase nao estao disponiveis, registra a nota."""
    ws.cell(row=1, column=1, value=titulo).font = Font(name="Calibri", bold=True, size=12)
    nota = ws.cell(
        row=3, column=1,
        value=f"Dados Supabase indisponiveis neste ciclo. Motivo: {motivo}",
    )
    nota.font = _ITALIC_FONT
    ws.column_dimensions["A"].width = 120


# ---------------------------------------------------------------------------
# Aba: Historico Comissoes
# ---------------------------------------------------------------------------

_HIST_HEADERS = [
    "nome", "cargo", "processo", "numero_pc", "codigo_cliente",
    "tipo", "tipo_pagamento", "documento", "nf_extraida", "linha_negocio",
    "status_processo", "mes_apuracao", "ano_apuracao",
    "valor_documento", "valor_processo",
    "tcmp", "fcmp_rampa", "fcmp_aplicado", "fcmp_considerado", "fcmp_modo",
    "comissao_potencial", "comissao_adiantada", "comissao_total",
    "status_faturamento_completo", "status_pagamento_completo", "reconciliado",
    "ac_snapshot_json", "af_snapshot_json",
    "tcmp_detalhes_json", "fcmp_detalhes_json",
    "created_at",
]


def escrever_aud_historico_comissoes(
    ws,
    historicos: Optional[Iterable[HistoricoComissao]],
    *,
    indisponivel_motivo: Optional[str] = None,
) -> None:
    ws.title = "Aud. Historico Comissoes"

    if indisponivel_motivo is not None:
        _write_unavailable_note(ws, "Aud. Historico Comissoes", indisponivel_motivo)
        return

    lista = list(historicos or [])
    _write_header(ws, 1, _HIST_HEADERS)

    if not lista:
        nota = ws.cell(row=2, column=1, value="Nenhum historico encontrado para esta GL.")
        nota.font = _ITALIC_FONT
        _auto_width(ws, 2, len(_HIST_HEADERS))
        return

    money_cols = {_HIST_HEADERS.index(c) + 1 for c in (
        "valor_documento", "valor_processo",
        "comissao_potencial", "comissao_adiantada", "comissao_total",
    )}
    factor_cols = {_HIST_HEADERS.index(c) + 1 for c in (
        "tcmp", "fcmp_rampa", "fcmp_aplicado", "fcmp_considerado",
    )}
    bool_cols = {_HIST_HEADERS.index(c) + 1 for c in (
        "status_faturamento_completo", "status_pagamento_completo", "reconciliado",
    )}
    json_cols = {_HIST_HEADERS.index(c) + 1 for c in (
        "ac_snapshot_json", "af_snapshot_json",
        "tcmp_detalhes_json", "fcmp_detalhes_json",
    )}
    datetime_cols = {_HIST_HEADERS.index("created_at") + 1}

    row = 2
    for h in lista:
        for col_idx, attr in enumerate(_HIST_HEADERS, start=1):
            value = getattr(h, attr, None)
            if col_idx in json_cols:
                _render_json_cell(ws, row, col_idx, value)
                continue
            if col_idx in bool_cols:
                cell = ws.cell(row=row, column=col_idx, value=_bool_label(value))
                continue
            if col_idx in datetime_cols:
                dt_val = _fmt_datetime(value)
                cell = ws.cell(row=row, column=col_idx, value=dt_val)
                if dt_val is not None:
                    cell.number_format = _DATETIME_FMT
                continue
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx in money_cols:
                cell.number_format = _MONEY_FMT
            elif col_idx in factor_cols:
                cell.number_format = _FACTOR_FMT
        row += 1

    _add_table(ws, 1, row - 1, len(_HIST_HEADERS), f"AudHist_{id(ws)%100000}")
    _auto_width(ws, row - 1, len(_HIST_HEADERS))
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Aba: Processos Pai
# ---------------------------------------------------------------------------

_PROC_PAI_HEADERS = [
    "numero_pc", "codigo_cliente", "processo",
    "is_processo_pai", "status_faturado", "status_pago",
    "mes_referencia", "ano_referencia", "created_at",
]


def escrever_aud_processos_pai(
    ws,
    vinculos: Optional[Iterable[HistoricoProcessoPai]],
    *,
    indisponivel_motivo: Optional[str] = None,
) -> None:
    ws.title = "Aud. Processos Pai"

    if indisponivel_motivo is not None:
        _write_unavailable_note(ws, "Aud. Processos Pai", indisponivel_motivo)
        return

    lista = list(vinculos or [])
    _write_header(ws, 1, _PROC_PAI_HEADERS)

    if not lista:
        nota = ws.cell(row=2, column=1, value="Nenhum vinculo de Processo Pai encontrado.")
        nota.font = _ITALIC_FONT
        _auto_width(ws, 2, len(_PROC_PAI_HEADERS))
        return

    bool_cols = {_PROC_PAI_HEADERS.index(c) + 1 for c in (
        "is_processo_pai", "status_faturado", "status_pago",
    )}
    datetime_cols = {_PROC_PAI_HEADERS.index("created_at") + 1}

    row = 2
    for v in lista:
        for col_idx, attr in enumerate(_PROC_PAI_HEADERS, start=1):
            value = getattr(v, attr, None)
            if col_idx in bool_cols:
                ws.cell(row=row, column=col_idx, value=_bool_label(value))
                continue
            if col_idx in datetime_cols:
                dt_val = _fmt_datetime(value)
                cell = ws.cell(row=row, column=col_idx, value=dt_val)
                if dt_val is not None:
                    cell.number_format = _DATETIME_FMT
                continue
            ws.cell(row=row, column=col_idx, value=value)
        row += 1

    _add_table(ws, 1, row - 1, len(_PROC_PAI_HEADERS), f"AudPai_{id(ws)%100000}")
    _auto_width(ws, row - 1, len(_PROC_PAI_HEADERS))
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Aba: Pagamentos Pai
# ---------------------------------------------------------------------------

_PAG_PAI_HEADERS = [
    "numero_pc", "codigo_cliente", "processo", "numero_nf", "documento",
    "situacao_codigo", "situacao_texto",
    "dt_prorrogacao", "data_baixa",
    "valor_documento", "mes_referencia", "ano_referencia", "created_at",
]


def escrever_aud_pagamentos_pai(
    ws,
    pagamentos: Optional[Iterable[HistoricoPagamentoProcessoPai]],
    *,
    indisponivel_motivo: Optional[str] = None,
) -> None:
    ws.title = "Aud. Pagamentos Pai"

    if indisponivel_motivo is not None:
        _write_unavailable_note(ws, "Aud. Pagamentos Pai", indisponivel_motivo)
        return

    lista = list(pagamentos or [])
    _write_header(ws, 1, _PAG_PAI_HEADERS)

    if not lista:
        nota = ws.cell(row=2, column=1, value="Nenhum pagamento de Processo Pai encontrado.")
        nota.font = _ITALIC_FONT
        _auto_width(ws, 2, len(_PAG_PAI_HEADERS))
        return

    money_cols = {_PAG_PAI_HEADERS.index("valor_documento") + 1}
    date_cols = {_PAG_PAI_HEADERS.index(c) + 1 for c in ("dt_prorrogacao", "data_baixa")}
    datetime_cols = {_PAG_PAI_HEADERS.index("created_at") + 1}

    row = 2
    for p in lista:
        for col_idx, attr in enumerate(_PAG_PAI_HEADERS, start=1):
            value = getattr(p, attr, None)
            if col_idx in date_cols:
                dt_val = _fmt_datetime(value)
                cell = ws.cell(row=row, column=col_idx, value=dt_val)
                if dt_val is not None:
                    cell.number_format = _DATE_FMT
                continue
            if col_idx in datetime_cols:
                dt_val = _fmt_datetime(value)
                cell = ws.cell(row=row, column=col_idx, value=dt_val)
                if dt_val is not None:
                    cell.number_format = _DATETIME_FMT
                continue
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx in money_cols:
                cell.number_format = _MONEY_FMT
        row += 1

    _add_table(ws, 1, row - 1, len(_PAG_PAI_HEADERS), f"AudPag_{id(ws)%100000}")
    _auto_width(ws, row - 1, len(_PAG_PAI_HEADERS))
    ws.freeze_panes = "A2"
