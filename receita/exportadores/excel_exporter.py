"""
receita/exportadores/excel_exporter.py — Exportação Excel por GL.

Gera recebimento_<nome>_MM_AAAA.xlsx com 9 abas:
    1. Resumo                       — consolidado por Linha de Negócio + conciliação do líquido
    2. Adiantamentos                — docs AF de processos ainda não totalmente faturados
    3. Pagamentos Regulares         — docs AF já faturados, sem reconciliação neste ciclo
    4. Reconciliação                — ajustes cross-month + estornos por devolução
    5. FCMP Processos               — métricas por processo (todos da competência)
    6. Detalhamento TCMP-FCMP       — auditoria item a item de TCMP e FCMP
    7. Aud. Historico Comissoes     — espelho da tabela Supabase (toda a historia da GL)
    8. Aud. Processos Pai           — espelho da tabela Supabase dos Pais relacionados
    9. Aud. Pagamentos Pai          — espelho da tabela Supabase das parcelas dos Pais

API pública
-----------
gerar_por_gl(comissao_result, reconciliacao_result, estornos_result,
             tcmp_result, fcmp_por_gl, saida_dir, mes, ano,
             status_por_processo_pai=None, historicos_por_gl=None) → List[str]
"""

from __future__ import annotations

import os
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from receita.exportadores import audit_sheets


# ---------------------------------------------------------------------------
# Constantes de formatação (espelham scripts/excel_export.py)
# ---------------------------------------------------------------------------

_MONEY_FMT = '#,##0.00'
_FACTOR_FMT = '0.0000'
_TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
_HEADER_FONT = Font(name="Calibri", bold=True, size=13)
_BOLD_FONT = Font(name="Calibri", bold=True)
_NORMAL_FONT = Font(name="Calibri", size=11)
_HEADER_FILL = PatternFill("solid", fgColor="2E75B6")
_HEADER_FONT_WHITE = Font(name="Calibri", bold=True, size=11, color="FFFFFF")


def _sanitize(nome: str) -> str:
    return (
        unicodedata.normalize("NFKD", nome)
        .encode("ascii", "ignore")
        .decode()
        .replace(" ", "_")
    )


def _nome_arquivo(nome_gl: str, mes: int, ano: int) -> str:
    return f"recebimento_{_sanitize(nome_gl)}_{mes:02d}_{ano}.xlsx"


def _auto_width(ws, start_row: int, end_row: int, num_cols: int):
    for col_idx in range(1, num_cols + 1):
        max_len = 10
        for row_idx in range(start_row, end_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)


def _add_table(ws, start_row: int, end_row: int, num_cols: int, table_name: str):
    ref = f"A{start_row}:{get_column_letter(num_cols)}{end_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = _TABLE_STYLE
    ws.add_table(tbl)


def _fmt_money(v: Any) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _write_header_row(ws, row: int, headers: List[str]):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = _HEADER_FONT_WHITE
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Aba 1: Resumo
# ---------------------------------------------------------------------------

def _aba_resumo(
    ws,
    itens_gl: List[Any],
    gl_nome: str,
    mes: int,
    ano: int,
    reconciliacao_result: Any = None,
    estornos_result: Any = None,
):
    ws.title = "Resumo"
    ws.cell(row=1, column=1, value=f"{gl_nome} — Recebimento {mes:02d}/{ano}").font = _HEADER_FONT
    ws.cell(row=2, column=1, value="Comissões por Linha de Negócio — regulares/adiantamentos pagos na competência; reconciliação e estornos entram como ajustes separados.").font = Font(name="Calibri", italic=True, color="666666")

    headers = ["Linha de Negócio", "Qtd Docs", "Valor Total (R$)", "Comissão Potencial (R$)",
               "FCMP Calculado", "Comissão (R$)"]
    _write_header_row(ws, 4, headers)

    adiantamentos, regulares = _classificar_itens_gl(itens_gl, gl_nome, reconciliacao_result)
    itens_ativos = adiantamentos + regulares

    grupos: Dict[str, Dict] = defaultdict(lambda: {"qtd": 0, "valor": 0.0, "potencial": 0.0, "final": 0.0, "fcmp": []})
    for item in itens_ativos:
        linha = getattr(item, "linha_negocio", "") or "—"
        g = grupos[linha]
        g["qtd"] += 1
        g["valor"] += _fmt_money(getattr(item, "valor_documento", 0))
        g["potencial"] += _fmt_money(getattr(item, "comissao_potencial", 0))
        g["final"] += _fmt_money(getattr(item, "comissao_final", 0))
        g["fcmp"].append(_fmt_money(getattr(item, "fcmp_aplicado", 1.0)))

    row = 5
    for linha, g in sorted(grupos.items()):
        fcmp_med = sum(g["fcmp"]) / len(g["fcmp"]) if g["fcmp"] else 1.0
        ws.cell(row=row, column=1, value=linha)
        ws.cell(row=row, column=2, value=g["qtd"])
        ws.cell(row=row, column=3, value=g["valor"]).number_format = _MONEY_FMT
        ws.cell(row=row, column=4, value=g["potencial"]).number_format = _MONEY_FMT
        ws.cell(row=row, column=5, value=fcmp_med).number_format = _FACTOR_FMT
        ws.cell(row=row, column=6, value=g["final"]).number_format = _MONEY_FMT
        row += 1

    if row > 5:
        _add_table(ws, 4, row - 1, len(headers), f"Resumo_{_sanitize(gl_nome)[:20]}")
        _auto_width(ws, 4, row - 1, len(headers))

    # ── Conciliação do Líquido ─────────────────────────────────────────────
    total_adiant = sum(_fmt_money(getattr(i, "comissao_final", 0)) for i in adiantamentos)
    total_regular = sum(_fmt_money(getattr(i, "comissao_final", 0)) for i in regulares)

    ajustes_recon = 0.0
    if reconciliacao_result is not None:
        for item in (getattr(reconciliacao_result, "itens", None) or []):
            if getattr(item, "gl_nome", "") == gl_nome:
                ajustes_recon += _fmt_money(getattr(item, "ajuste", 0))

    estornos_valor = 0.0
    if estornos_result is not None:
        for item in (getattr(estornos_result, "itens", None) or []):
            if getattr(item, "gl_nome", "") == gl_nome:
                estornos_valor += _fmt_money(getattr(item, "estorno", 0))

    row += 1
    titulo_cell = ws.cell(row=row, column=1, value="Conciliação do Líquido a Pagar")
    titulo_cell.font = _HEADER_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    _write_header_row(ws, row, ["Item", "Valor (R$)"])
    row += 1

    linhas_conc = [
        ("Comissão de Adiantamentos", total_adiant),
        ("Comissão de Pagamentos Regulares", total_regular),
        ("(+/-) Ajustes de Reconciliação", ajustes_recon),
        ("(+) Estornos por Devolução (negativos)", estornos_valor),
    ]
    total_liquido = total_adiant + total_regular + ajustes_recon + estornos_valor
    for descricao, valor in linhas_conc:
        ws.cell(row=row, column=1, value=descricao)
        vc = ws.cell(row=row, column=2, value=valor)
        vc.number_format = _MONEY_FMT
        row += 1

    total_desc = ws.cell(row=row, column=1, value="TOTAL LÍQUIDO A PAGAR")
    total_desc.font = _BOLD_FONT
    total_desc.fill = _TOTAL_FILL
    total_cell = ws.cell(row=row, column=2, value=total_liquido)
    total_cell.number_format = _MONEY_FMT
    total_cell.font = _BOLD_FONT
    total_cell.fill = _TOTAL_FILL


# ---------------------------------------------------------------------------
# Helpers de classificacao por cenario
# ---------------------------------------------------------------------------

def _status_label(value: Any) -> str:
    """Converte bool/None para 'Sim' / 'Nao' / '—' (status pagamento pode ser None)."""
    if value is None:
        return "—"
    return "Sim" if value else "Não"


def _processos_reconciliados_gl(
    gl_nome: str,
    reconciliacao_result: Any,
) -> set:
    """Retorna set de (numero_pc, codigo_cliente, processo) absorvidos pela aba Reconciliacao."""
    if reconciliacao_result is None:
        return set()
    itens = getattr(reconciliacao_result, "itens", None) or []
    chaves: set = set()
    for item in itens:
        if getattr(item, "gl_nome", "") != gl_nome:
            continue
        pc = getattr(item, "numero_pc", "") or ""
        cli = getattr(item, "codigo_cliente", "") or ""
        detalhes = getattr(item, "detalhes_historicos", None) or []
        if detalhes:
            for d in detalhes:
                p = str(d.get("processo", "") or "").strip().upper()
                if p:
                    chaves.add((str(pc).strip().upper(), str(cli).strip().upper(), p))
        proc_item = str(getattr(item, "processo", "") or "").strip().upper()
        if proc_item:
            chaves.add((str(pc).strip().upper(), str(cli).strip().upper(), proc_item))
    return chaves


def _classificar_itens_gl(
    itens_gl: List[Any],
    gl_nome: str,
    reconciliacao_result: Any,
) -> Tuple[List[Any], List[Any]]:
    """Divide itens do GL em (adiantamentos, regulares).

    Regra:
      1. tipo_pagamento == ADIANTAMENTO OU provisorio == True -> adiantamentos.
      2. Caso contrario -> regulares.

    Itens de processos reconciliados neste ciclo NÃO são suprimidos — a comissão
    de competência permanece a pagar no mês (FCMP_considerado=1,0) e o Ajuste
    da Reconciliação corrige o total para o FCMP real.
    """
    adiantamentos: List[Any] = []
    regulares: List[Any] = []
    for item in itens_gl:
        tipo = str(getattr(item, "tipo_pagamento", "") or "").strip().upper()
        provisorio = bool(getattr(item, "provisorio", False))
        if tipo == "ADIANTAMENTO" or provisorio:
            adiantamentos.append(item)
        else:
            regulares.append(item)
    return adiantamentos, regulares


# ---------------------------------------------------------------------------
# Aba 2: Adiantamentos
# ---------------------------------------------------------------------------

def _aba_adiantamentos(ws, itens: List[Any]):
    """Adiantamentos — processos ainda nao totalmente faturados. FCMP sempre 1.0."""
    ws.title = "Adiantamentos"
    headers = [
        "Processo", "Documento", "NF Extraída",
        "Linha de Negócio", "Status Processo",
        "Valor Documento (R$)", "TCMP",
        "Fat. Completo?", "Pag. Completo?",
        "FCMP Modo", "Comissão Adiantada (R$)",
    ]

    nota = ws.cell(row=1, column=1, value=(
        "Adiantamentos — Comissão = Valor × TCMP. FCMP = 1,0 (provisório). "
        "O FCMP real e o ajuste definitivo entram na aba 'Reconciliação' quando o Processo Pai fechar."
    ))
    nota.font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    _write_header_row(ws, 3, headers)

    if not itens:
        ws.cell(row=4, column=1, value="Sem adiantamentos neste ciclo.").font = Font(italic=True, color="888888")
        _auto_width(ws, 3, 4, len(headers))
        return

    row_idx = 4
    for item in itens:
        values = [
            getattr(item, "processo", ""),
            getattr(item, "documento", ""),
            getattr(item, "nf_extraida", ""),
            getattr(item, "linha_negocio", ""),
            getattr(item, "status_processo", ""),
            _fmt_money(getattr(item, "valor_documento", 0)),
            _fmt_money(getattr(item, "tcmp", 0)),
            _status_label(getattr(item, "status_faturamento_completo", None)),
            _status_label(getattr(item, "status_pagamento_completo", None)),
            getattr(item, "fcmp_modo", "") or "PROVISÓRIO",
            _fmt_money(getattr(item, "comissao_final", 0)),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col == 6 or col == 11:
                cell.number_format = _MONEY_FMT
            elif col == 7:
                cell.number_format = _FACTOR_FMT
        row_idx += 1

    _add_table(ws, 3, row_idx - 1, len(headers), f"Adiantamentos_{row_idx}")
    _auto_width(ws, 3, row_idx - 1, len(headers))
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Aba 3: Pagamentos Regulares
# ---------------------------------------------------------------------------

def _aba_pagamentos_regulares(ws, itens: List[Any]):
    """Pagamentos regulares — processos faturados, aguardando reconciliacao do Pai."""
    ws.title = "Pagamentos Regulares"
    headers = [
        "Processo", "Documento", "NF Extraída", "Tipo Pagamento",
        "Linha de Negócio", "Status Processo",
        "Valor Documento (R$)", "TCMP",
        "FCMP Rampa", "FCMP Calculado (pós-fat.)", "FCMP Modo",
        "Fat. Completo?", "Pag. Completo?", "Aguarda Reconciliação",
        "Comissão a Pagar (R$)",
    ]

    nota = ws.cell(row=1, column=1, value=(
        "Pagamentos Regulares — Comissão a Pagar na competência usa FCMP = 1,0. "
        "O 'FCMP Calculado (pós-fat.)' é o valor real calculado após o faturamento; "
        "entrará no ajuste de Reconciliação quando o Processo Pai fechar."
    ))
    nota.font = Font(italic=True, color="666666")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    _write_header_row(ws, 3, headers)

    if not itens:
        ws.cell(row=4, column=1, value="Sem pagamentos regulares neste ciclo.").font = Font(italic=True, color="888888")
        _auto_width(ws, 3, 4, len(headers))
        return

    row_idx = 4
    for item in itens:
        status_pag = getattr(item, "status_pagamento_completo", None)
        aguarda = "Sim" if status_pag is True else "Não"
        values = [
            getattr(item, "processo", ""),
            getattr(item, "documento", ""),
            getattr(item, "nf_extraida", ""),
            getattr(item, "tipo_pagamento", "") or "REGULAR",
            getattr(item, "linha_negocio", ""),
            getattr(item, "status_processo", ""),
            _fmt_money(getattr(item, "valor_documento", 0)),
            _fmt_money(getattr(item, "tcmp", 0)),
            _fmt_money(getattr(item, "fcmp_rampa", 1)),
            _fmt_money(getattr(item, "fcmp_aplicado", 1)),
            getattr(item, "fcmp_modo", ""),
            _status_label(getattr(item, "status_faturamento_completo", None)),
            _status_label(status_pag),
            aguarda,
            _fmt_money(getattr(item, "comissao_final", 0)),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col in (7, 15):
                cell.number_format = _MONEY_FMT
            elif col in (8, 9, 10):
                cell.number_format = _FACTOR_FMT
        row_idx += 1

    _add_table(ws, 3, row_idx - 1, len(headers), f"Regulares_{row_idx}")
    _auto_width(ws, 3, row_idx - 1, len(headers))
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Aba 3: FCMP por Processo
# ---------------------------------------------------------------------------

def _aba_fcmp_processos(
    ws,
    itens_gl: List[Any],
    tcmp_result: Any,
    fcmp_result_gl: Any,
    gl_nome: str = "",
    reconciliacao_result: Any = None,
):
    ws.title = "FCMP Processos"
    headers = [
        "Processo", "Status", "Qtd Itens AC", "Valor Faturado (R$)",
        "Qtd Docs AF", "Valor Recebido (R$)", "TCMP", "FCMP Rampa",
        "FCMP Calculado", "Modo", "Provisório", "Comissão Total (R$)"
    ]
    _write_header_row(ws, 1, headers)

    # Agrupar docs por processo — inclui processos reconciliados (aba diagnóstica).
    docs_por_proc: Dict[str, list] = defaultdict(list)
    for item in itens_gl:
        proc = getattr(item, "processo", "")
        if not proc:
            continue
        docs_por_proc[proc].append(item)

    tcmp_map: Dict[str, float] = {}
    if tcmp_result:
        if hasattr(tcmp_result, "tcmp_por_processo"):
            tcmp_map = tcmp_result.tcmp_por_processo
        elif isinstance(tcmp_result, dict):
            tcmp_map = tcmp_result.get("tcmp_por_processo", {})

    fcmp_map: Dict[str, Any] = {}
    if fcmp_result_gl:
        if hasattr(fcmp_result_gl, "fcmp_por_processo"):
            fcmp_map = fcmp_result_gl.fcmp_por_processo

    row_idx = 2
    for processo in sorted(docs_por_proc.keys()):
        docs = docs_por_proc[processo]
        fcmp_proc = fcmp_map.get(processo)

        tcmp = tcmp_map.get(processo, docs[0].tcmp if docs else 0.0)
        fcmp_r = fcmp_proc.fcmp_rampa if fcmp_proc else (docs[0].fcmp_rampa if docs else 1.0)
        fcmp_a = fcmp_proc.fcmp_aplicado if fcmp_proc else (docs[0].fcmp_aplicado if docs else 1.0)
        modo = fcmp_proc.modo if fcmp_proc else (docs[0].fcmp_modo if docs else "RAMPA")
        prov = fcmp_proc.provisorio if fcmp_proc else (docs[0].provisorio if docs else False)
        num_itens_ac = fcmp_proc.num_itens if fcmp_proc else 0
        valor_fat = fcmp_proc.valor_faturado if fcmp_proc else 0.0
        valor_rec = sum(_fmt_money(getattr(d, "valor_documento", 0)) for d in docs)
        comissao_total = sum(_fmt_money(getattr(d, "comissao_final", 0)) for d in docs)
        status = docs[0].status_processo if docs else ""

        values = [
            processo, status, num_itens_ac, valor_fat,
            len(docs), valor_rec, tcmp, fcmp_r,
            fcmp_a, modo, "Sim" if prov else "Não", comissao_total
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            if col in (4, 6, 12):
                cell.number_format = _MONEY_FMT
            elif col in (7, 8, 9):
                cell.number_format = _FACTOR_FMT
        row_idx += 1

    if row_idx > 2:
        _add_table(ws, 1, row_idx - 1, len(headers), f"FCMPProc_{row_idx}")
        _auto_width(ws, 1, row_idx - 1, len(headers))


# ---------------------------------------------------------------------------
# Aba 5: Detalhamento TCMP/FCMP
# ---------------------------------------------------------------------------

_HIERARQUIA_LABELS = ["L1 Linha", "L2 Grupo", "L3 Subgrupo", "L4 Tipo Mercad.", "L5 Fabricante", "L6 Aplicação"]

_SECTION_FILL = PatternFill("solid", fgColor="1F4E79")
_SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
_SUBSECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
_SUBSECTION_FONT = Font(name="Calibri", bold=True, size=11)
_TOTAL_FILL = PatternFill("solid", fgColor="FFF2CC")
_TOTAL_FONT = Font(name="Calibri", bold=True)
_PROV_FILL = PatternFill("solid", fgColor="F8CBAD")


def _write_section_header(ws, row: int, text: str, num_cols: int) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _SECTION_FONT
    cell.fill = _SECTION_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[row].height = 20
    if num_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    return row + 1


def _write_subsection_header(ws, row: int, text: str, num_cols: int) -> int:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _SUBSECTION_FONT
    cell.fill = _SUBSECTION_FILL
    ws.row_dimensions[row].height = 16
    if num_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    return row + 1


_COMPONENTE_LABELS = {
    "faturamento_linha": "Faturamento (hierarquia)",
    "faturamento_individual": "Faturamento (individual)",
    "conversao_linha": "Conversão (hierarquia)",
    "conversao_individual": "Conversão (individual)",
    "rentabilidade": "Rentabilidade",
    "retencao_clientes": "Retenção de Clientes",
    "meta_fornecedor_1": "Meta Fornecedor 1",
    "meta_fornecedor_2": "Meta Fornecedor 2",
}


def _label_componente(nome: str) -> str:
    return _COMPONENTE_LABELS.get(nome, nome)


def _aba_fcmp_componentes_bloco(
    ws,
    row: int,
    itens_fcmp: List[Dict],
    num_cols: int,
    gl_rule_label: Optional[Dict[str, str]] = None,
) -> int:
    """Renderiza C.1 — componentes de FC por hierarquia única encontrada no processo.

    Cada entrada exibe: peso, realizado, meta, atingimento, atingimento cap,
    contribuição. O FC rampa dessa combinação aparece como linha-total.
    """
    vistos: Dict[str, Dict[str, Any]] = {}
    for det in itens_fcmp:
        fc_detail = det.get("fc_item_detail") or {}
        hier_key = fc_detail.get("hierarquia_key", "")
        if hier_key and hier_key not in vistos:
            vistos[hier_key] = fc_detail

    if not vistos:
        return row

    row = _write_subsection_header(
        ws, row,
        "C.1 — Como o FC foi calculado para cada hierarquia (Realizado vs Meta × Peso)",
        num_cols,
    )

    for hier_key, fc_detail in vistos.items():
        colaborador = fc_detail.get("colaborador", "")
        cargo = fc_detail.get("cargo", "")
        fc_rampa_item = float(fc_detail.get("fc_rampa", 0.0))
        fc_final_item = float(fc_detail.get("fc_final", 0.0))
        modo_item = fc_detail.get("modo", "RAMPA")
        componentes = fc_detail.get("componentes", []) or []

        rule_display = (gl_rule_label or {}).get(hier_key, hier_key)
        titulo = ws.cell(
            row=row, column=1,
            value=(
                f"GL: {colaborador}  |  Cargo: {cargo}  |  Regra GL: {rule_display}  "
                f"|  FC Rampa: {fc_rampa_item:.4f}  |  FC Final: {fc_final_item:.4f}  ({modo_item})"
            ),
        )
        titulo.font = _BOLD_FONT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        row += 1

        comp_headers = [
            "Componente", "Peso", "Realizado", "Meta",
            "Atingimento", "Atingimento (cap)", "Contribuição",
        ]
        _write_header_row(ws, row, comp_headers)
        start = row
        row += 1

        if not componentes:
            ws.cell(row=row, column=1,
                    value="Nenhum componente com peso > 0 para este (cargo, colaborador).")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
            row += 1
        else:
            soma_contrib = 0.0
            for comp in componentes:
                peso = float(comp.get("peso", 0.0))
                realizado = float(comp.get("realizado", 0.0))
                meta = float(comp.get("meta", 0.0))
                ating = float(comp.get("atingimento", 0.0))
                ating_cap = float(comp.get("atingimento_cap", 0.0))
                contrib = float(comp.get("contribuicao", 0.0))
                soma_contrib += contrib

                vals = [
                    _label_componente(str(comp.get("nome", ""))),
                    peso, realizado, meta, ating, ating_cap, contrib,
                ]
                for col, v in enumerate(vals, start=1):
                    c = ws.cell(row=row, column=col, value=v)
                    if col == 2:  # Peso
                        c.number_format = '0.00%'
                    elif col in (3, 4):  # Realizado, Meta
                        c.number_format = _MONEY_FMT
                    elif col in (5, 6, 7):  # Atingimento / cap / Contribuição
                        c.number_format = _FACTOR_FMT
                row += 1

            _add_table(ws, start, row - 1, len(comp_headers),
                       f"FCComp_{_sanitize(hier_key)[:10]}_{start}")

            formula = ws.cell(
                row=row, column=1,
                value=(
                    f"FC Rampa = Σ Contribuição = {soma_contrib:.4f}  "
                    f"(cap aplicado se > cap_fc_max)  →  FC Rampa final = {fc_rampa_item:.4f}"
                ),
            )
            formula.font = _TOTAL_FONT
            formula.fill = _TOTAL_FILL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
            row += 1

        row += 1  # espaço entre hierarquias

    return row


def _aba_fcmp_escada_bloco(
    ws,
    row: int,
    itens_fcmp: List[Dict],
    fcmp_rampa: float,
    fcmp_aplicado: float,
    num_cols: int,
) -> int:
    """Renderiza C.3 — escada do cargo usada para discretizar o FCMP Rampa."""
    escada_info: Optional[Dict[str, Any]] = None
    for det in itens_fcmp:
        fc_detail = det.get("fc_item_detail") or {}
        if fc_detail.get("modo") == "ESCADA":
            escada_info = fc_detail
            break

    row = _write_subsection_header(
        ws, row,
        f"C.3 — Escada do cargo (FCMP Rampa {fcmp_rampa:.4f} → FCMP Aplicado {fcmp_aplicado:.4f})",
        num_cols,
    )

    if not escada_info:
        ws.cell(row=row, column=1,
                value="Escada aplicada ao FCMP do processo (config do cargo — ver configuracoes_comissoes).")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        return row + 1

    num_degraus = escada_info.get("escada_num_degraus")
    piso = escada_info.get("escada_piso")

    try:
        from scripts.fc_calculator import gerar_degraus_escada
        if num_degraus and piso is not None:
            degraus = gerar_degraus_escada(int(num_degraus), float(piso), None)
        else:
            degraus = []
    except Exception:
        degraus = []

    if not degraus:
        ws.cell(row=row, column=1,
                value=(
                    f"Modo ESCADA — piso={piso}, num_degraus={num_degraus}. "
                    "Detalhamento dos degraus indisponível."
                ))
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
        return row + 1

    # Menor degrau >= fcmp_rampa (round-up, igual a _aplicar_escada)
    if fcmp_rampa >= 1.0:
        idx_aplicado = len(degraus) - 1
    else:
        idx_aplicado = len(degraus) - 1
        for k, d in enumerate(degraus):
            if d >= fcmp_rampa:
                idx_aplicado = k
                break

    _write_header_row(ws, row, ["Degrau", "Valor", "Aplicado?"])
    start = row
    row += 1
    for i, d in enumerate(degraus):
        ws.cell(row=row, column=1, value=i)
        c2 = ws.cell(row=row, column=2, value=d)
        c2.number_format = _FACTOR_FMT
        aplic = "← APLICADO" if i == idx_aplicado else ""
        c3 = ws.cell(row=row, column=3, value=aplic)
        if i == idx_aplicado:
            c3.font = _TOTAL_FONT
            c3.fill = _TOTAL_FILL
        row += 1

    _add_table(ws, start, row - 1, 3, f"Escada_{start}")

    formula_esc = ws.cell(
        row=row, column=1,
        value=(
            f"FCMP Aplicado = {fcmp_aplicado:.4f}  "
            f"(FCMP Rampa {fcmp_rampa:.4f} mapeado ao menor degrau ≥ rampa)"
        ),
    )
    formula_esc.font = _TOTAL_FONT
    formula_esc.fill = _TOTAL_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    return row + 1


def _aba_detalhamento_tcmp_fcmp(
    ws,
    itens_gl: List[Any],
    tcmp_result: Any,
    fcmp_result_gl: Any,
):
    """Aba 5 — Detalhamento passo a passo de TCMP e FCMP por processo.

    Para cada processo que gerou comissão neste GL:
    - Seção A: documentos AF com a fórmula de cálculo da comissão
    - Seção B: breakdown item a item do TCMP
    - Seção C: breakdown item a item do FCMP (ou nota provisória)
    """
    ws.title = "Detalhamento TCMP-FCMP"
    MAX_COLS = 11  # largura total usada para merge de cabeçalhos

    # Extrair mapa TCMP detalhes
    tcmp_detalhes: Dict[str, List[Dict]] = {}
    if tcmp_result:
        if hasattr(tcmp_result, "detalhes"):
            tcmp_detalhes = tcmp_result.detalhes or {}
        elif isinstance(tcmp_result, dict):
            tcmp_detalhes = tcmp_result.get("detalhes", {})

    # Extrair mapa FCMP detalhes
    fcmp_detalhes: Dict[str, List[Dict]] = {}
    fcmp_map: Dict[str, Any] = {}
    if fcmp_result_gl:
        if hasattr(fcmp_result_gl, "detalhes"):
            fcmp_detalhes = fcmp_result_gl.detalhes or {}
        elif isinstance(fcmp_result_gl, dict):
            fcmp_detalhes = fcmp_result_gl.get("detalhes", {})
        if hasattr(fcmp_result_gl, "fcmp_por_processo"):
            fcmp_map = fcmp_result_gl.fcmp_por_processo
        elif isinstance(fcmp_result_gl, dict):
            fcmp_map = fcmp_result_gl.get("fcmp_por_processo", {})

    # Agrupar documentos por processo (ordem estável)
    docs_por_proc: Dict[str, list] = defaultdict(list)
    proc_order: list = []
    seen: set = set()
    for item in itens_gl:
        proc = getattr(item, "processo", "") or ""
        if proc and proc not in seen:
            proc_order.append(proc)
            seen.add(proc)
        if proc:
            docs_por_proc[proc].append(item)

    row = 1
    ws.cell(row=row, column=1, value="Detalhamento TCMP/FCMP — Auditoria de Cálculo").font = _HEADER_FONT
    row += 2

    for processo in proc_order:
        docs = docs_por_proc[processo]
        primeiro = docs[0]
        status_proc = getattr(primeiro, "status_processo", "") or ""
        linha_neg = getattr(primeiro, "linha_negocio", "") or ""

        # ── Cabeçalho do processo ──────────────────────────────────────────
        row = _write_section_header(
            ws, row,
            f"Processo {processo}   |   Status: {status_proc}   |   Linha: {linha_neg}",
            MAX_COLS,
        )

        # ── Seção A: Documentos ───────────────────────────────────────────
        row = _write_subsection_header(ws, row, "A — Documentos e fórmula de comissão", MAX_COLS)

        doc_headers = [
            "Documento", "Tipo", "Valor Doc. (R$)",
            "TCMP", "FCMP Calculado (pós-fat.)",
            "Fórmula (FCMP efetivo)", "Comissão Final (R$)",
        ]
        _write_header_row(ws, row, doc_headers)
        doc_header_row = row
        row += 1

        for item in docs:
            tipo = getattr(item, "tipo_pagamento", "") or ""
            valor = _fmt_money(getattr(item, "valor_documento", 0))
            tcmp_val = _fmt_money(getattr(item, "tcmp", 0))
            fcmp_aplic = _fmt_money(getattr(item, "fcmp_aplicado", 1))
            fcmp_cons = _fmt_money(getattr(item, "fcmp_considerado", 1))
            comissao = _fmt_money(getattr(item, "comissao_final", 0))

            # Fórmula textual: usa fcmp_considerado (= 1,0 até reconciliação)
            if tipo == "ADIANTAMENTO":
                fc_display = "1,0 (adiantamento)"
            else:
                fc_display = f"{fcmp_cons:.4f} (FCMP efetivo; calculado = {fcmp_aplic:.4f})"
            formula_txt = (
                f"{valor:,.2f} × {tcmp_val:.4%} × {fc_display}"
                f" = {comissao:,.2f}"
            )

            values = [
                getattr(item, "documento", ""),
                tipo,
                valor,
                tcmp_val,
                fcmp_aplic,
                formula_txt,
                comissao,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                if col == 3:
                    cell.number_format = _MONEY_FMT
                elif col in (4, 5):
                    cell.number_format = _FACTOR_FMT
                elif col == 7:
                    cell.number_format = _MONEY_FMT
            row += 1

        if len(docs) > 0:
            _add_table(ws, doc_header_row, row - 1, len(doc_headers),
                       f"DocsDet_{_sanitize(processo)[:15]}_{doc_header_row}")

        row += 1  # espaço

        # ── Seção B: TCMP breakdown ───────────────────────────────────────
        tcmp_val_proc = 0.0
        if tcmp_result:
            if hasattr(tcmp_result, "tcmp_por_processo"):
                tcmp_val_proc = tcmp_result.tcmp_por_processo.get(processo, 0.0)
            elif isinstance(tcmp_result, dict):
                tcmp_val_proc = tcmp_result.get("tcmp_por_processo", {}).get(processo, 0.0)

        row = _write_subsection_header(
            ws, row,
            f"B — Como foi calculado o TCMP = {tcmp_val_proc:.4%}",
            MAX_COLS,
        )

        nota_tcmp = ws.cell(
            row=row, column=1,
            value=(
                "Para cada item, a regra de GL mais específica (por hierarquia) define a Taxa Efetiva = "
                "Fatia Cargo (%) × Taxa Máx Cargo (%) / 100. "
                "Contribuição = Valor Peso × Taxa Efetiva. "
                "TCMP do processo = Σ Contribuição / Σ Valor Peso."
            ),
        )
        nota_tcmp.font = Font(italic=True, color="555555")
        nota_tcmp.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COLS)
        row += 1

        itens_tcmp = tcmp_detalhes.get(processo, [])
        if itens_tcmp:
            tcmp_item_headers = _HIERARQUIA_LABELS + [
                "Valor Peso (R$)",
                "Taxa Máx Cargo (%)",
                "Fatia Cargo (%)",
                "Taxa Efetiva",
                "Contribuição",
            ]
            _write_header_row(ws, row, tcmp_item_headers)
            tcmp_table_start = row
            row += 1

            soma_valor_tcmp = 0.0
            soma_contrib_tcmp = 0.0
            for detalhe in itens_tcmp:
                hier = detalhe.get("hierarquia") or ("",) * 6
                taxa = float(detalhe.get("taxa", 0))
                val_item = float(detalhe.get("valor_item", 0))
                contrib = float(detalhe.get("contribuicao", 0))
                taxa_max = float(detalhe.get("taxa_rateio_maximo_pct", 0))
                fatia = float(detalhe.get("fatia_cargo_pct", 0))
                soma_valor_tcmp += val_item
                soma_contrib_tcmp += contrib

                row_vals = list(hier) + [val_item, taxa_max, fatia, taxa, contrib]
                for col, val in enumerate(row_vals, start=1):
                    cell = ws.cell(row=row, column=col, value=val)
                    if col == 7:  # Valor Peso
                        cell.number_format = _MONEY_FMT
                    elif col in (8, 9):  # Taxa Máx Cargo (%), Fatia Cargo (%)
                        cell.number_format = '0.00'
                    elif col in (10, 11):  # Taxa Efetiva, Contribuição
                        cell.number_format = _FACTOR_FMT
                row += 1

            _add_table(ws, tcmp_table_start, row - 1, len(tcmp_item_headers),
                       f"TCMP_{_sanitize(processo)[:15]}_{tcmp_table_start}")

            # Linha de total e fórmula
            total_cell = ws.cell(row=row, column=7, value=soma_valor_tcmp)
            total_cell.number_format = _MONEY_FMT
            total_cell.font = _TOTAL_FONT
            total_cell.fill = _TOTAL_FILL
            contrib_cell = ws.cell(row=row, column=11, value=soma_contrib_tcmp)
            contrib_cell.number_format = _FACTOR_FMT
            contrib_cell.font = _TOTAL_FONT
            contrib_cell.fill = _TOTAL_FILL
            row += 1

            formula_cell = ws.cell(
                row=row, column=1,
                value=(
                    f"TCMP = Σ Contribuição / Σ Valor Peso  =  "
                    f"{soma_contrib_tcmp:.6f} / {soma_valor_tcmp:,.2f}  =  {tcmp_val_proc:.4%}"
                ),
            )
            formula_cell.font = _TOTAL_FONT
            formula_cell.fill = _TOTAL_FILL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COLS)
        else:
            note = ws.cell(row=row, column=1,
                           value="Sem detalhamento disponível para este processo.")
            note.font = Font(italic=True, color="888888")
        row += 2  # espaço

        # ── Seção C: FCMP breakdown ───────────────────────────────────────
        fcmp_proc = fcmp_map.get(processo) if fcmp_map else None
        fcmp_rampa = float(getattr(fcmp_proc, "fcmp_rampa", 1.0) if fcmp_proc else 1.0)
        fcmp_aplicado = float(getattr(fcmp_proc, "fcmp_aplicado", 1.0) if fcmp_proc else 1.0)
        fcmp_modo = str(getattr(fcmp_proc, "modo", "RAMPA") if fcmp_proc else "RAMPA")
        is_provisorio = bool(getattr(fcmp_proc, "provisorio", False) if fcmp_proc else True)

        row = _write_subsection_header(
            ws, row,
            f"C — Como foi calculado o FCMP = {fcmp_aplicado:.4f}  (modo: {fcmp_modo})",
            MAX_COLS,
        )

        if is_provisorio:
            prov_cell = ws.cell(
                row=row, column=1,
                value=(
                    "FCMP = 1,0 (PROVISÓRIO) — processo ainda não FATURADO. "
                    "O FC real será recalculado e o ajuste aplicado na reconciliação "
                    "quando o Processo Pai fechar."
                ),
            )
            prov_cell.font = Font(italic=True, color="C00000", bold=True)
            prov_cell.fill = _PROV_FILL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COLS)
            row += 1
        else:
            itens_fcmp = fcmp_detalhes.get(processo, [])
            if itens_fcmp:
                # Mapa: hierarquia_key do item → label da regra GL (do TCMP detalhe)
                _gl_rule_label: Dict[str, str] = {}
                for _t in tcmp_detalhes.get(processo, []):
                    _item_key = "/".join(h for h in (_t.get("hierarquia") or ()) if h)
                    _gl_hier = "/".join(h for h in (_t.get("gl_hierarquia") or ()) if h)
                    if _item_key and _gl_hier:
                        _gl_rule_label[_item_key] = _gl_hier

                # C.1 — Componentes do FC agrupados por hierarquia única
                row = _aba_fcmp_componentes_bloco(ws, row, itens_fcmp, MAX_COLS, _gl_rule_label)

                # C.2 — FCMP item a item (média ponderada por valor)
                row = _write_subsection_header(
                    ws, row,
                    "C.2 — Média ponderada dos FC por item (FCMP Rampa)",
                    MAX_COLS,
                )
                fcmp_item_headers = _HIERARQUIA_LABELS + ["Valor (R$)", "FC Item", "Contribuição"]
                _write_header_row(ws, row, fcmp_item_headers)
                fcmp_table_start = row
                row += 1

                soma_valor_fcmp = 0.0
                soma_contrib_fcmp = 0.0
                for detalhe in itens_fcmp:
                    hier = detalhe.get("hierarquia") or ("",) * 6
                    fc_item = float(detalhe.get("fc_item", 1.0))
                    val_item = float(detalhe.get("valor_item", 0))
                    contrib = float(detalhe.get("contribuicao", 0))
                    soma_valor_fcmp += val_item
                    soma_contrib_fcmp += contrib

                    row_vals = list(hier) + [val_item, fc_item, contrib]
                    for col, val in enumerate(row_vals, start=1):
                        cell = ws.cell(row=row, column=col, value=val)
                        if col == 7:  # Valor
                            cell.number_format = _MONEY_FMT
                        elif col in (8, 9):  # FC Item, Contribuição
                            cell.number_format = _FACTOR_FMT
                    row += 1

                _add_table(ws, fcmp_table_start, row - 1, len(fcmp_item_headers),
                           f"FCMP_{_sanitize(processo)[:15]}_{fcmp_table_start}")

                # Linha de total
                tv = ws.cell(row=row, column=7, value=soma_valor_fcmp)
                tv.number_format = _MONEY_FMT
                tv.font = _TOTAL_FONT
                tv.fill = _TOTAL_FILL
                tc = ws.cell(row=row, column=9, value=soma_contrib_fcmp)
                tc.number_format = _FACTOR_FMT
                tc.font = _TOTAL_FONT
                tc.fill = _TOTAL_FILL
                row += 1

                # Fórmula FCMP Rampa
                formula_rampa = ws.cell(
                    row=row, column=1,
                    value=(
                        f"FCMP Rampa = Σ Contribuição / Σ Valor  =  "
                        f"{soma_contrib_fcmp:.6f} / {soma_valor_fcmp:,.2f}  =  {fcmp_rampa:.4f}"
                    ),
                )
                formula_rampa.font = _TOTAL_FONT
                formula_rampa.fill = _TOTAL_FILL
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COLS)
                row += 1

                # C.3 — Escada (se aplicável)
                if fcmp_modo == "ESCADA":
                    row = _aba_fcmp_escada_bloco(
                        ws, row, itens_fcmp, fcmp_rampa, fcmp_aplicado, MAX_COLS,
                    )
            else:
                note = ws.cell(row=row, column=1,
                               value="Sem detalhamento item a item disponível para o FCMP deste processo.")
                note.font = Font(italic=True, color="888888")
                row += 1

        row += 2  # espaço entre processos

    # Ajuste de largura das colunas
    _auto_width(ws, 1, row - 1, MAX_COLS)


# ---------------------------------------------------------------------------
# Aba 4: Reconciliação e Estornos
# ---------------------------------------------------------------------------

def _aba_reconciliacao_estornos(ws, gl_nome: str, reconciliacao_result: Any, estornos_result: Any):
    ws.title = "Reconciliação"
    MAX_COLS = 7
    row = 1

    # ── Reconciliação ─────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="Reconciliação de Adiantamentos").font = _HEADER_FONT
    row += 1

    itens_rec = []
    if reconciliacao_result:
        itens_raw = getattr(reconciliacao_result, "itens", None) or []
        itens_rec = [i for i in itens_raw if getattr(i, "gl_nome", "") == gl_nome]

    if not itens_rec:
        ws.cell(row=row, column=1, value="Sem reconciliações neste ciclo.").font = Font(italic=True, color="888888")
        row += 2
    else:
        # Resumo geral
        rec_headers = ["Processo", "PC", "Cliente", "Comissão Adiantada (R$)", "FCMP Real", "Ajuste (R$)", "Históricos"]
        _write_header_row(ws, row, rec_headers)
        rec_start = row
        row += 1

        for item in itens_rec:
            values = [
                getattr(item, "processo", ""),
                getattr(item, "numero_pc", ""),
                getattr(item, "codigo_cliente", ""),
                _fmt_money(getattr(item, "comissao_adiantada", 0)),
                _fmt_money(getattr(item, "fcmp_real", 1)),
                _fmt_money(getattr(item, "ajuste", 0)),
                int(getattr(item, "historicos_considerados", 0)),
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                if col in (4, 6):
                    cell.number_format = _MONEY_FMT
                elif col == 5:
                    cell.number_format = _FACTOR_FMT
            row += 1

        if row > rec_start + 1:
            _add_table(ws, rec_start, row - 1, len(rec_headers), "Reconciliacao")
        row += 1

        # ── Detalhamento passo a passo por reconciliação ──────────────
        row = _write_section_header(ws, row, "Detalhamento passo a passo das Reconciliações", MAX_COLS)
        row += 1

        for item_idx, item in enumerate(itens_rec):
            processo = getattr(item, "processo", "")
            pc = getattr(item, "numero_pc", "")
            cliente = getattr(item, "codigo_cliente", "")
            comissao_total = _fmt_money(getattr(item, "comissao_adiantada", 0))
            fcmp_real = _fmt_money(getattr(item, "fcmp_real", 1))
            ajuste = _fmt_money(getattr(item, "ajuste", 0))
            detalhes = getattr(item, "detalhes_historicos", []) or []

            processos_lista = sorted({
                str(d.get("processo", "")) for d in detalhes if d.get("processo")
            }) if detalhes else [processo]
            row = _write_subsection_header(
                ws, row,
                f"Reconciliação — PC: {pc}  |  Cliente: {cliente}  |  Processos: {', '.join(processos_lista)}",
                MAX_COLS,
            )

            if detalhes:
                det_headers = [
                    "Processo", "Documento", "Tipo", "Mês/Ano Apuração",
                    "Comissão Adiantada (R$)", "FCMP Aplicado",
                    "Contribuição Ponderada (R$)", "Entra no FCMP Real?",
                ]
                _write_header_row(ws, row, det_headers)
                det_start = row
                row += 1

                soma_comissao_total = 0.0
                soma_comissao_fcmp = 0.0
                soma_ponderada_fcmp = 0.0

                for det in detalhes:
                    d_processo = det.get("processo", "")
                    d_doc = det.get("documento", "")
                    d_tipo = str(det.get("tipo_pagamento", "") or "").strip().upper() or "REGULAR"
                    d_mes = det.get("mes_apuracao", 0)
                    d_ano = det.get("ano_apuracao", 0)
                    d_comissao = float(det.get("comissao_adiantada", 0))
                    d_fcmp = float(det.get("fcmp_aplicado", 1))
                    d_contrib = float(det.get("contribuicao_ponderada", 0))
                    # Padrão: adiantamentos (FCMP provisório=1,0) são excluídos do FCMP Real.
                    d_entra = bool(det.get("contribui_fcmp_real", d_tipo != "ADIANTAMENTO"))

                    soma_comissao_total += d_comissao
                    if d_entra:
                        soma_comissao_fcmp += d_comissao
                        soma_ponderada_fcmp += d_contrib

                    det_values = [
                        d_processo,
                        d_doc,
                        d_tipo,
                        f"{d_mes:02d}/{d_ano}" if d_mes and d_ano else "",
                        d_comissao,
                        d_fcmp,
                        d_contrib if d_entra else 0.0,
                        "Sim" if d_entra else "Não (provisório)",
                    ]
                    for col, val in enumerate(det_values, start=1):
                        cell = ws.cell(row=row, column=col, value=val)
                        if col in (5, 7):
                            cell.number_format = _MONEY_FMT
                        elif col == 6:
                            cell.number_format = _FACTOR_FMT
                    row += 1

                tbl_name = f"RecDet_{item_idx}_{det_start}"
                _add_table(ws, det_start, row - 1, len(det_headers), tbl_name)

                # Linha de totais
                tc1 = ws.cell(row=row, column=5, value=soma_comissao_total)
                tc1.number_format = _MONEY_FMT
                tc1.font = _TOTAL_FONT
                tc1.fill = _TOTAL_FILL
                tc2 = ws.cell(row=row, column=7, value=soma_ponderada_fcmp)
                tc2.number_format = _MONEY_FMT
                tc2.font = _TOTAL_FONT
                tc2.fill = _TOTAL_FILL
                row += 1

                # Fórmula FCMP Real (apenas rows pós-faturamento)
                if soma_comissao_fcmp > 0:
                    formula_fcmp_txt = (
                        f"FCMP Real = Σ Contribuição Ponderada (pós-fat.) / Σ Comissão Adiantada (pós-fat.)  =  "
                        f"{soma_ponderada_fcmp:,.6f} / {soma_comissao_fcmp:,.2f}  =  {fcmp_real:.4f}"
                    )
                else:
                    formula_fcmp_txt = (
                        f"FCMP Real = 1,0000 (sem rows pós-faturamento; só adiantamentos provisórios)"
                    )
                formula_fcmp = ws.cell(row=row, column=1, value=formula_fcmp_txt)
                formula_fcmp.font = _TOTAL_FONT
                formula_fcmp.fill = _TOTAL_FILL
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COLS)
                row += 1

                # Fórmula Ajuste (base = total de comissão adiantada, incluindo provisórios)
                formula_ajuste = ws.cell(
                    row=row, column=1,
                    value=(
                        f"Ajuste = Comissão Adiantada Total × (FCMP Real - 1,0)  =  "
                        f"{soma_comissao_total:,.2f} × ({fcmp_real:.4f} - 1,0000)  =  {ajuste:,.2f}"
                    ),
                )
                formula_ajuste.font = _TOTAL_FONT
                formula_ajuste.fill = _TOTAL_FILL
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_COLS)
                row += 1
            else:
                ws.cell(
                    row=row, column=1,
                    value="Sem detalhamento disponível (históricos não capturados).",
                ).font = Font(italic=True, color="888888")
                row += 1

            row += 1  # espaço entre reconciliações

    # ── Estornos ──────────────────────────────────────────────────────────
    row += 1
    ws.cell(row=row, column=1, value="Estornos por Devolução").font = _HEADER_FONT
    row += 1
    est_headers = ["Processo", "NF Origem", "Valor Devolvido (R$)", "Comissão Base (R$)", "Estorno (R$)"]
    _write_header_row(ws, row, est_headers)
    row += 1
    est_start = row

    itens_est = []
    if estornos_result:
        itens_raw = getattr(estornos_result, "itens", None) or []
        itens_est = [i for i in itens_raw if getattr(i, "gl_nome", "") == gl_nome]

    for item in itens_est:
        values = [
            getattr(item, "processo", ""),
            getattr(item, "nf_origem", ""),
            _fmt_money(getattr(item, "valor_devolvido", 0)),
            _fmt_money(getattr(item, "comissao_base", 0)),
            _fmt_money(getattr(item, "estorno", 0)),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            if col in (3, 4, 5):
                cell.number_format = _MONEY_FMT
        row += 1

    if not itens_est:
        ws.cell(row=row, column=1, value="Sem estornos neste ciclo.").font = Font(italic=True, color="888888")
    elif row > est_start:
        _add_table(ws, est_start - 1, row - 1, len(est_headers), "Estornos")

    _auto_width(ws, 1, row, MAX_COLS)


# ---------------------------------------------------------------------------
# Função pública
# ---------------------------------------------------------------------------

def _enriquecer_itens_com_status_pai(
    itens: List[Any],
    status_por_processo_pai: Optional[Dict[str, Dict]],
    processo_para_pai: Optional[Dict[str, Optional[str]]],
) -> None:
    """Injeta nos items (in-place) os campos do Processo Pai.

    Campos adicionados: numero_pc, codigo_cliente,
    status_faturamento_completo, status_pagamento_completo.
    Se ja existirem no item, nao sobrescreve valores truthy.
    """
    if not status_por_processo_pai:
        return
    mapa_proc = processo_para_pai or {}
    for item in itens:
        proc = str(getattr(item, "processo", "") or "").strip().upper()
        if not proc:
            continue
        chave = mapa_proc.get(proc)
        if not chave:
            # fallback: tentar via numero_pc+codigo_cliente ja presentes
            pc_exist = str(getattr(item, "numero_pc", "") or "").strip().upper()
            cli_exist = str(getattr(item, "codigo_cliente", "") or "").strip().upper()
            if pc_exist and cli_exist:
                chave = f"{pc_exist}|{cli_exist}"
            else:
                continue
        status = status_por_processo_pai.get(chave)
        if not status:
            continue
        if not getattr(item, "numero_pc", "") and status.get("numero_pc"):
            try:
                setattr(item, "numero_pc", status["numero_pc"])
            except Exception:
                pass
        if not getattr(item, "codigo_cliente", "") and status.get("codigo_cliente"):
            try:
                setattr(item, "codigo_cliente", status["codigo_cliente"])
            except Exception:
                pass
        if getattr(item, "status_faturamento_completo", None) in (None, False):
            try:
                setattr(item, "status_faturamento_completo",
                        bool(status.get("status_faturamento_completo", False)))
            except Exception:
                pass
        if getattr(item, "status_pagamento_completo", None) is None:
            try:
                setattr(item, "status_pagamento_completo",
                        status.get("status_pagamento_completo"))
            except Exception:
                pass


# ---------------------------------------------------------------------------
# AUDITORIA — nova aba: Itens AC+AF com coloração por comissão do GL
# ---------------------------------------------------------------------------

_AUDIT_AC_COLS: List[str] = [
    "Código Produto", "Descrição Produto", "Qtde Solicitada", "Qtde Atendida",
    "Preço Unitário", "Operação", "Processo", "Numero NF", "Status Processo",
    "Dt Entrada", "Status da NF", "Data Aceite", "Dt Aprovação", "Dt Emissão",
    "Valor Orçado", "Valor Realizado", "Consultor Interno", "Representante-pedido",
    "Gerente Comercial-Pedido", "Aplicação Mat./Serv.", "Cliente", "Nome Cliente",
    "Cidade", "UF",
]

# Rótulos de cabeçalho AF → nome da coluna no df_af_mapeado (ou None p/ valor direto)
_AUDIT_AF_COLS: List[Tuple[str, str]] = [
    ("Documento", "Documento"),
    ("Cliente 2", "Cliente"),
    ("Dt. Vencimento", "Dt. Vencimento"),
    ("Situação", "Situação"),
    ("Valor Líquido", "Valor Líquido"),
    ("Data de Baixa", "Data de Baixa"),
    ("Tipo de Baixa", "Tipo de Baixa"),
]

_AUDIT_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_AUDIT_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def _audit_norm(v: Any) -> str:
    if v is None:
        return ""
    try:
        import pandas as _pd
        if _pd.isna(v):
            return ""
    except Exception:
        pass
    return str(v).strip().upper()


def _audit_cell_value(v: Any) -> Any:
    if v is None:
        return ""
    try:
        import pandas as _pd
        if _pd.isna(v):
            return ""
        if hasattr(v, "to_pydatetime"):
            return v.to_pydatetime()
    except Exception:
        pass
    return v


def _aba_auditoria_ac_af(
    ws,
    gl_nome: str,
    itens_gl: List[Any],
    df_ac_full: Any,
    df_af_mapeado: Any,
    mes: int,
    ano: int,
):
    """Aba extra — itens da AC do mês + documentos AF vinculados, com cor por comissão.

    Verde: (Processo, Documento) gerou comissão para o GL.
    Vermelho: não gerou.
    Ordenação: verdes primeiro, vermelhos depois; dentro de cada grupo por Documento.
    """
    import pandas as pd

    ws.title = "Itens AC+AF (Auditoria)"

    if df_ac_full is None or getattr(df_ac_full, "empty", True):
        ws.cell(row=1, column=1, value="AC indisponível — aba de auditoria vazia.")
        return

    # Filtrar AC pelo mês/ano selecionado via Dt Emissão
    df_ac = df_ac_full
    if "Dt Emissão" in df_ac_full.columns:
        try:
            dt = pd.to_datetime(df_ac_full["Dt Emissão"], errors="coerce")
            mask = (dt.dt.month == int(mes)) & (dt.dt.year == int(ano))
            df_ac = df_ac_full[mask].copy()
        except Exception:
            df_ac = df_ac_full.copy()

    chaves_comissao = {
        (_audit_norm(getattr(i, "processo", "")), _audit_norm(getattr(i, "documento", "")))
        for i in itens_gl
    }

    # Indexar AF por processo_ac (resultado do mapeamento)
    af_por_processo: Dict[str, List[Any]] = defaultdict(list)
    if df_af_mapeado is not None and not getattr(df_af_mapeado, "empty", True):
        proc_af_col = None
        for candidato in ["processo_ac", "processo", "Processo"]:
            if candidato in df_af_mapeado.columns:
                proc_af_col = candidato
                break
        if proc_af_col:
            for _, af_row in df_af_mapeado.iterrows():
                proc_norm = _audit_norm(af_row.get(proc_af_col))
                if proc_norm:
                    af_por_processo[proc_norm].append(af_row)

    headers = _AUDIT_AC_COLS + [label for label, _ in _AUDIT_AF_COLS]
    for col_idx, h in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=h).font = Font(bold=True)

    # Montar linhas: AC_item × AF_docs (um por doc); AC_item sem AF → 1 linha
    linhas: List[Tuple[bool, str, Any, Optional[Any]]] = []  # (is_green, documento_sort, ac_row, af_row|None)
    for _, ac_row in df_ac.iterrows():
        proc_norm = _audit_norm(ac_row.get("Processo"))
        af_rows = af_por_processo.get(proc_norm, [])
        if not af_rows:
            linhas.append((False, "", ac_row, None))
            continue
        for af_row in af_rows:
            doc_norm = _audit_norm(af_row.get("Documento"))
            is_green = (proc_norm, doc_norm) in chaves_comissao
            linhas.append((is_green, doc_norm, ac_row, af_row))

    # Ordenação: verdes primeiro; dentro de cada grupo por Documento
    linhas.sort(key=lambda t: (0 if t[0] else 1, t[1]))

    start_data = 2
    ac_cols_presentes = [c for c in _AUDIT_AC_COLS if c in df_ac.columns]
    af_cols_presentes = []
    if df_af_mapeado is not None and not getattr(df_af_mapeado, "empty", True):
        af_cols_presentes = [src for _, src in _AUDIT_AF_COLS if src in df_af_mapeado.columns]

    for row_idx, (is_green, _, ac_row, af_row) in enumerate(linhas, start=start_data):
        fill = _AUDIT_GREEN_FILL if is_green else _AUDIT_RED_FILL
        col_idx = 1
        for col_name in _AUDIT_AC_COLS:
            val = ac_row.get(col_name, "") if col_name in ac_cols_presentes else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=_audit_cell_value(val))
            cell.fill = fill
            if col_name in ("Valor Orçado", "Valor Realizado", "Preço Unitário"):
                cell.number_format = _MONEY_FMT
            col_idx += 1
        for _label, src in _AUDIT_AF_COLS:
            if af_row is not None and src in af_cols_presentes:
                val = af_row.get(src, "")
            else:
                val = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=_audit_cell_value(val))
            cell.fill = fill
            if src == "Valor Líquido":
                cell.number_format = _MONEY_FMT
            col_idx += 1

    end_row = start_data + len(linhas) - 1
    if end_row >= start_data:
        ws.freeze_panes = "A2"
        _auto_width(ws, 1, end_row, len(headers))


def gerar_por_gl(
    comissao_result: Any,
    reconciliacao_result: Any,
    estornos_result: Any,
    tcmp_result: Any,
    fcmp_por_gl: Any,
    saida_dir: str,
    mes: int,
    ano: int,
    status_por_processo_pai: Optional[Dict[str, Dict]] = None,
    processo_para_pai: Optional[Dict[str, Optional[str]]] = None,
    historicos_por_gl: Optional[Dict[str, Dict[str, Any]]] = None,
    df_ac_full: Any = None,
    df_af_mapeado: Any = None,
) -> List[str]:
    """Gera arquivo .xlsx de recebimento para cada GL com comissões.

    Args:
        comissao_result: ComissaoResult ou dict.
        reconciliacao_result: ReconciliacaoResult ou dict.
        estornos_result: EstornosResult ou dict.
        tcmp_result: TCMPResult ou dict.
        fcmp_por_gl: {gl_nome: FCMPResult} ou dict.
        saida_dir: Diretório de saída.
        mes: Mês de apuração.
        ano: Ano de apuração.
        status_por_processo_pai: {f"{pc}|{cli}": status_dict} com
            status_faturamento_completo / status_pagamento_completo (opcional).
        processo_para_pai: {processo: "pc|cli"} para enriquecer itens (opcional).
        historicos_por_gl: {gl_nome: {"historicos": [...], "vinculos": [...],
            "pagamentos": [...], "erro": str|None}} vindo das consultas Supabase
            para as abas de auditoria (opcional).

    Returns:
        Lista de caminhos dos arquivos .xlsx gerados.
    """
    os.makedirs(saida_dir, exist_ok=True)
    arquivos: List[str] = []

    # Suporte a objeto ou dict
    if hasattr(comissao_result, "itens"):
        todos_itens = comissao_result.itens
    else:
        itens_raw = comissao_result.get("itens", []) if isinstance(comissao_result, dict) else []
        from types import SimpleNamespace
        todos_itens = [SimpleNamespace(**i) for i in itens_raw]

    # Enriquecer itens com status do Processo Pai (se disponivel)
    _enriquecer_itens_com_status_pai(todos_itens, status_por_processo_pai, processo_para_pai)

    # Agrupar por GL
    por_gl: Dict[str, list] = defaultdict(list)
    for item in todos_itens:
        gl = getattr(item, "gl_nome", "")
        if gl:
            por_gl[gl].append(item)

    historicos_por_gl = historicos_por_gl or {}

    for gl_nome, itens_gl in por_gl.items():
        # Obter FCMPResult do GL
        fcmp_result_gl = None
        if fcmp_por_gl:
            fcmp_raw = fcmp_por_gl.get(gl_nome) if isinstance(fcmp_por_gl, dict) else None
            if fcmp_raw is not None:
                if hasattr(fcmp_raw, "fcmp_por_processo"):
                    fcmp_result_gl = fcmp_raw
                else:
                    from receita.schemas.calculo import FCMPProcesso, FCMPResult
                    fcmp_por_proc = {}
                    for p, fp_raw in fcmp_raw.get("fcmp_por_processo", {}).items():
                        fcmp_por_proc[p] = FCMPProcesso(**fp_raw)
                    fcmp_result_gl = FCMPResult(gl_nome=gl_nome, fcmp_por_processo=fcmp_por_proc)

        adiantamentos, regulares = _classificar_itens_gl(itens_gl, gl_nome, reconciliacao_result)

        wb = openpyxl.Workbook()

        # Aba 1: Resumo
        ws1 = wb.active
        _aba_resumo(ws1, itens_gl, gl_nome, mes, ano,
                    reconciliacao_result=reconciliacao_result,
                    estornos_result=estornos_result)

        # Aba 2: Adiantamentos
        ws_adiant = wb.create_sheet()
        _aba_adiantamentos(ws_adiant, adiantamentos)

        # Aba 3: Pagamentos Regulares
        ws_reg = wb.create_sheet()
        _aba_pagamentos_regulares(ws_reg, regulares)

        # Aba 4: Reconciliacao + Estornos
        ws_recon = wb.create_sheet()
        _aba_reconciliacao_estornos(ws_recon, gl_nome, reconciliacao_result, estornos_result)

        # Aba 5: FCMP Processos (todos os processos da competência)
        ws_fcmp = wb.create_sheet()
        _aba_fcmp_processos(
            ws_fcmp, itens_gl, tcmp_result, fcmp_result_gl,
            gl_nome=gl_nome, reconciliacao_result=reconciliacao_result,
        )

        # Aba 6: Detalhamento TCMP-FCMP
        ws_det = wb.create_sheet()
        _aba_detalhamento_tcmp_fcmp(ws_det, itens_gl, tcmp_result, fcmp_result_gl)

        # Abas 7-9: Auditoria Supabase
        dados_aud = historicos_por_gl.get(gl_nome, {}) or {}
        erro_aud = dados_aud.get("erro")

        ws_aud_hist = wb.create_sheet()
        audit_sheets.escrever_aud_historico_comissoes(
            ws_aud_hist,
            dados_aud.get("historicos"),
            indisponivel_motivo=erro_aud,
        )

        ws_aud_pai = wb.create_sheet()
        audit_sheets.escrever_aud_processos_pai(
            ws_aud_pai,
            dados_aud.get("vinculos"),
            indisponivel_motivo=erro_aud,
        )

        ws_aud_pag = wb.create_sheet()
        audit_sheets.escrever_aud_pagamentos_pai(
            ws_aud_pag,
            dados_aud.get("pagamentos"),
            indisponivel_motivo=erro_aud,
        )

        # Aba 10: Auditoria Itens AC + AF (somente se dataframes fornecidos)
        if df_ac_full is not None:
            ws_aud_acaf = wb.create_sheet()
            _aba_auditoria_ac_af(
                ws_aud_acaf, gl_nome, itens_gl,
                df_ac_full, df_af_mapeado, mes, ano,
            )

        nome_arq = _nome_arquivo(gl_nome, mes, ano)
        caminho = Path(saida_dir) / nome_arq
        wb.save(str(caminho))
        arquivos.append(str(caminho))

    return arquivos
